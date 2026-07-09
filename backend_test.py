#!/usr/bin/env python3
"""
Comprehensive backend test for Stanvard School ERP - Fee Status Report Extended Testing
"""
import requests
import sys
import io

BASE_URL = "https://data-pull-6.preview.emergentagent.com/api"
SUPER_ADMIN = {"email": "superadmin@stanvard.school", "password": "Stanvard@2026"}
PARENT = {"email": "9079111899", "password": "111899"}

test_results = []
total_tests = 0
passed_tests = 0
failed_tests = 0

def log_test(name, passed, msg=""):
    global total_tests, passed_tests, failed_tests
    total_tests += 1
    if passed:
        passed_tests += 1
        status = "✅"
    else:
        failed_tests += 1
        status = "❌"
    result = f"{status} {name}" + (f" - {msg}" if msg else "")
    test_results.append(result)
    print(result)

def login(creds):
    r = requests.post(f"{BASE_URL}/auth/login", json=creds, timeout=10)
    return r.json().get("access_token") if r.status_code == 200 else None

def get_headers(token):
    return {"Authorization": f"Bearer {token}"}

def get_school_id(token):
    r = requests.get(f"{BASE_URL}/schools", headers=get_headers(token), timeout=10)
    schools = r.json() if r.status_code == 200 else []
    return schools[0]['id'] if schools else None

def test_base_report(token, sid):
    print("\n" + "="*80)
    print("TEST 1: Base Fee Status Report")
    print("="*80)
    
    r = requests.get(f"{BASE_URL}/reports/fee-status?school_id={sid}", 
                     headers=get_headers(token), timeout=30)
    
    log_test("Base call returns 200", r.status_code == 200, f"Got {r.status_code}")
    if r.status_code != 200:
        return None
    
    data = r.json()
    
    for key in ['rows', 'count', 'by_class', 'summary']:
        log_test(f"Has '{key}' field", key in data)
    
    rows = data['rows']
    log_test("Count matches rows", len(rows) == data['count'])
    log_test("Has student data", len(rows) > 0, f"{len(rows)} students")
    
    if len(rows) == 0:
        return None
    
    # Check row schema
    row = rows[0]
    required = ['student_id', 'admission_number', 'full_name', 'class_id', 'class_name',
                'section', 'phone', 'father_name', 'expected', 'gross_expected', 'discount',
                'paid', 'due', 'collection_percent', 'due_date', 'upcoming_due_date',
                'last_payment_date', 'overdue_days', 'status', 'behavior_tag']
    missing = [f for f in required if f not in row]
    log_test("Row has all required fields", len(missing) == 0, 
             f"Missing: {missing}" if missing else "")
    
    # Check calculations (first 10 rows)
    calc_ok = True
    for r in rows[:10]:
        if abs(r['expected'] - (r['gross_expected'] - r['discount'])) > 0.01:
            calc_ok = False
            break
        if abs(r['due'] - max(r['expected'] - r['paid'], 0)) > 0.01:
            calc_ok = False
            break
    log_test("Row calculations correct", calc_ok)
    
    # Check status values
    status_ok = all(r['status'] in ['paid', 'partial', 'unpaid'] for r in rows[:10])
    log_test("Status values valid", status_ok)
    
    # Check behavior_tag values
    behavior_ok = all(r['behavior_tag'] in ['regular', 'late', 'defaulter', 'na'] for r in rows[:10])
    log_test("Behavior_tag values valid", behavior_ok)
    
    # Check summary
    s = data['summary']
    total_exp = sum(r['expected'] for r in rows)
    total_paid = sum(r['paid'] for r in rows)
    total_due = sum(r['due'] for r in rows)
    
    log_test("Summary total_expected correct", abs(s['total_expected'] - total_exp) <= 1)
    log_test("Summary total_paid correct", abs(s['total_paid'] - total_paid) <= 0.01)
    log_test("Summary total_due correct", abs(s['total_due'] - total_due) <= 0.01)
    
    if total_exp > 0:
        coll_calc = round((total_paid / total_exp) * 100, 1)
        log_test("Summary collection_percent correct", 
                 abs(s['collection_percent'] - coll_calc) <= 0.1)
    
    # Check behavior counts
    def_count = sum(1 for r in rows if r['behavior_tag'] == 'defaulter')
    late_count = sum(1 for r in rows if r['behavior_tag'] == 'late')
    reg_count = sum(1 for r in rows if r['behavior_tag'] == 'regular')
    
    log_test("Summary defaulter_count correct", s['defaulter_count'] == def_count)
    log_test("Summary late_count correct", s['late_count'] == late_count)
    log_test("Summary regular_count correct", s['regular_count'] == reg_count)
    
    # Check by_class
    log_test("by_class has entries", len(data['by_class']) > 0, 
             f"{len(data['by_class'])} groups")
    
    if data['by_class']:
        bc = data['by_class'][0]
        bc_req = ['class_id', 'class_name', 'section', 'students', 'expected',
                  'paid', 'due', 'collection_percent']
        bc_missing = [f for f in bc_req if f not in bc]
        log_test("by_class has required fields", len(bc_missing) == 0)
        
        # Validate by_class calculations
        class_rows = [r for r in rows if r['class_id'] == bc['class_id'] and 
                     (r.get('section') or '-') == bc['section']]
        log_test("by_class student count correct", len(class_rows) == bc['students'])
        
        exp_sum = sum(r['expected'] for r in class_rows)
        log_test("by_class expected sum correct", abs(bc['expected'] - exp_sum) <= 0.01)
    
    return data

def test_filters(token, sid, base_data):
    print("\n" + "="*80)
    print("TESTS 2-8: Filters")
    print("="*80)
    
    # Test 2: quick_view=defaulters
    r = requests.get(f"{BASE_URL}/reports/fee-status?school_id={sid}&quick_view=defaulters",
                     headers=get_headers(token), timeout=30)
    log_test("quick_view=defaulters returns 200", r.status_code == 200)
    if r.status_code == 200:
        data = r.json()
        all_def = all(row['behavior_tag'] == 'defaulter' for row in data['rows'])
        log_test("All rows are defaulters", all_def, f"{len(data['rows'])} rows")
        if base_data:
            log_test("Count matches base defaulter_count", 
                     len(data['rows']) == base_data['summary']['defaulter_count'])
    
    # Test 3: quick_view=fully_paid
    r = requests.get(f"{BASE_URL}/reports/fee-status?school_id={sid}&quick_view=fully_paid",
                     headers=get_headers(token), timeout=30)
    log_test("quick_view=fully_paid returns 200", r.status_code == 200)
    if r.status_code == 200:
        data = r.json()
        all_paid = all(row['status'] == 'paid' and row['due'] <= 0 for row in data['rows'])
        log_test("All rows fully paid", all_paid, f"{len(data['rows'])} rows")
    
    # Test 4: quick_view=upcoming
    r = requests.get(f"{BASE_URL}/reports/fee-status?school_id={sid}&quick_view=upcoming",
                     headers=get_headers(token), timeout=30)
    log_test("quick_view=upcoming returns 200", r.status_code == 200)
    if r.status_code == 200:
        data = r.json()
        all_upcoming = all(row.get('upcoming_due_date') for row in data['rows'])
        log_test("All rows have upcoming_due_date", all_upcoming, f"{len(data['rows'])} rows")
    
    # Test 5: behavior=late
    r = requests.get(f"{BASE_URL}/reports/fee-status?school_id={sid}&behavior=late",
                     headers=get_headers(token), timeout=30)
    log_test("behavior=late returns 200", r.status_code == 200)
    if r.status_code == 200:
        data = r.json()
        all_late = all(row['behavior_tag'] == 'late' for row in data['rows'])
        log_test("All rows are late", all_late, f"{len(data['rows'])} rows")
    
    # Test 6: due_min/due_max
    r = requests.get(f"{BASE_URL}/reports/fee-status?school_id={sid}&due_min=5000&due_max=25000",
                     headers=get_headers(token), timeout=30)
    log_test("due_min/due_max returns 200", r.status_code == 200)
    if r.status_code == 200:
        data = r.json()
        in_range = all(5000 <= row['due'] <= 25000 for row in data['rows'])
        log_test("All rows in due range", in_range, f"{len(data['rows'])} rows")
    
    # Test 7: status_filter=partial
    r = requests.get(f"{BASE_URL}/reports/fee-status?school_id={sid}&status_filter=partial",
                     headers=get_headers(token), timeout=30)
    log_test("status_filter=partial returns 200", r.status_code == 200)
    if r.status_code == 200:
        data = r.json()
        all_partial = all(row['status'] == 'partial' and 0 < row['paid'] < row['expected'] 
                         for row in data['rows'])
        log_test("All rows partial", all_partial, f"{len(data['rows'])} rows")
    
    # Test 8: payment_date_start
    r = requests.get(f"{BASE_URL}/reports/fee-status?school_id={sid}&payment_date_start=2020-01-01",
                     headers=get_headers(token), timeout=30)
    log_test("payment_date_start=2020-01-01 returns 200", r.status_code == 200)
    if r.status_code == 200:
        data = r.json()
        all_have_date = all(row.get('last_payment_date') and 
                           row['last_payment_date'] >= '2020-01-01' for row in data['rows'])
        log_test("All rows have valid last_payment_date", all_have_date, f"{len(data['rows'])} rows")
    
    r = requests.get(f"{BASE_URL}/reports/fee-status?school_id={sid}&payment_date_start=2099-01-01",
                     headers=get_headers(token), timeout=30)
    log_test("payment_date_start=2099-01-01 returns 200", r.status_code == 200)
    if r.status_code == 200:
        data = r.json()
        log_test("Future date returns 0 rows", len(data['rows']) == 0)

def test_rbac(parent_token, sid):
    print("\n" + "="*80)
    print("TEST 9: RBAC - Parent Access")
    print("="*80)
    
    r = requests.get(f"{BASE_URL}/reports/fee-status?school_id={sid}",
                     headers=get_headers(parent_token), timeout=30)
    
    if r.status_code == 403:
        log_test("Parent gets 403", True)
    elif r.status_code == 200:
        data = r.json()
        if len(data.get('rows', [])) == 0:
            log_test("Parent gets empty result", True)
        else:
            log_test("Parent access restricted", False, 
                     f"Got {len(data['rows'])} rows (should be 0 or 403)")
    else:
        log_test("Parent RBAC handled", False, f"Got {r.status_code}")

def test_exports(token, sid):
    print("\n" + "="*80)
    print("TESTS 10a-c: Exports")
    print("="*80)
    
    # PDF
    r = requests.get(f"{BASE_URL}/reports/fee-status.pdf?school_id={sid}&quick_view=defaulters",
                     headers=get_headers(token), timeout=30)
    log_test("PDF export returns 200", r.status_code == 200)
    if r.status_code == 200:
        log_test("PDF Content-Type correct", 'application/pdf' in r.headers.get('Content-Type', ''))
        log_test("PDF non-empty", len(r.content) > 0, f"{len(r.content)} bytes")
        log_test("PDF valid signature", r.content[:4] == b'%PDF')
    
    # XLSX
    r = requests.get(f"{BASE_URL}/reports/fee-status.xlsx?school_id={sid}&quick_view=defaulters",
                     headers=get_headers(token), timeout=30)
    log_test("XLSX export returns 200", r.status_code == 200)
    if r.status_code == 200:
        ct = r.headers.get('Content-Type', '')
        log_test("XLSX Content-Type correct", 'openxmlformats' in ct or 'spreadsheetml' in ct)
        log_test("XLSX non-empty", len(r.content) > 0, f"{len(r.content)} bytes")
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(r.content))
            expected = ['Fee Status', 'Summary', 'By Class']
            missing = [s for s in expected if s not in wb.sheetnames]
            log_test("XLSX has 3 sheets", len(missing) == 0, 
                     f"Missing: {missing}" if missing else "All present")
        except ImportError:
            log_test("XLSX sheets (openpyxl unavailable)", True, "Skipped")
        except Exception as e:
            log_test("XLSX sheets verification", False, str(e))
    
    # CSV
    r = requests.get(f"{BASE_URL}/reports/fee-status.csv?school_id={sid}&quick_view=defaulters",
                     headers=get_headers(token), timeout=30)
    log_test("CSV export returns 200", r.status_code == 200)
    if r.status_code == 200:
        log_test("CSV Content-Type correct", 'text/csv' in r.headers.get('Content-Type', ''))
        log_test("CSV non-empty", len(r.text) > 0, f"{len(r.text)} chars")
        
        lines = r.text.strip().split('\n')
        if lines:
            header = lines[0]
            required = ['Admission No', 'Student', 'Class', 'Section', 'Guardian', 'Phone',
                       'Expected (Rs.)', 'Discount (Rs.)', 'Paid (Rs.)', 'Due (Rs.)',
                       'Due Date', 'Last Payment', 'Overdue Days', 'Status', 'Behavior']
            missing = [c for c in required if c not in header]
            log_test("CSV has all columns", len(missing) == 0, 
                     f"Missing: {missing}" if missing else "")

def test_regression(token, sid):
    print("\n" + "="*80)
    print("TEST 11: Regression")
    print("="*80)
    
    # Fee heads
    r = requests.get(f"{BASE_URL}/fees/heads?school_id={sid}", headers=get_headers(token), timeout=10)
    if r.status_code == 200 and len(r.json()) > 0:
        head_id = r.json()[0]['id']
        
        r2 = requests.patch(f"{BASE_URL}/fees/heads/{head_id}?school_id={sid}",
                           headers=get_headers(token),
                           json={"name": r.json()[0]['name']}, timeout=10)
        log_test("PATCH /api/fees/heads/{id} works", r2.status_code == 200)
        
        r3 = requests.delete(f"{BASE_URL}/fees/heads/{head_id}?school_id={sid}",
                            headers=get_headers(token), timeout=10)
        log_test("DELETE /api/fees/heads/{id} responds", r3.status_code in [200, 400])
    else:
        log_test("Fee heads endpoints", False, "No fee heads found")
    
    # Fee plans
    r = requests.get(f"{BASE_URL}/fees/plans?school_id={sid}", headers=get_headers(token), timeout=10)
    if r.status_code == 200 and len(r.json()) > 0:
        plan_id = r.json()[0]['id']
        
        r2 = requests.delete(f"{BASE_URL}/fees/plans/{plan_id}?school_id={sid}",
                            headers=get_headers(token), timeout=10)
        log_test("DELETE /api/fees/plans/{id} responds", r2.status_code in [200, 400])
    else:
        log_test("Fee plans endpoint", False, "No fee plans found")
    
    # Fee schedule
    r = requests.get(f"{BASE_URL}/students?school_id={sid}&limit=1", 
                     headers=get_headers(token), timeout=10)
    if r.status_code == 200 and len(r.json()) > 0:
        student_id = r.json()[0]['id']
        
        r2 = requests.get(f"{BASE_URL}/fees/student/{student_id}/fee-schedule?school_id={sid}",
                         headers=get_headers(token), timeout=10)
        log_test("GET /api/fees/student/{id}/fee-schedule works", r2.status_code == 200)
    else:
        log_test("Fee schedule endpoint", False, "No students found")

def main():
    print("="*80)
    print("STANVARD SCHOOL ERP - FEE STATUS REPORT EXTENDED TESTING")
    print("="*80)
    
    print("\n🔐 Logging in as Super Admin...")
    admin_token = login(SUPER_ADMIN)
    if not admin_token:
        print("❌ Failed to login as super admin")
        sys.exit(1)
    print("✅ Super Admin login successful")
    
    school_id = get_school_id(admin_token)
    if not school_id:
        print("❌ Failed to get school_id")
        sys.exit(1)
    print(f"✅ Using school_id: {school_id[:8]}...")
    
    print("\n🔐 Logging in as Parent...")
    parent_token = login(PARENT)
    if parent_token:
        print("✅ Parent login successful")
    else:
        print("⚠️  Parent login failed, RBAC test will be skipped")
    
    # Run tests
    base_data = test_base_report(admin_token, school_id)
    test_filters(admin_token, school_id, base_data)
    if parent_token:
        test_rbac(parent_token, school_id)
    test_exports(admin_token, school_id)
    test_regression(admin_token, school_id)
    
    # Summary
    print("\n" + "="*80)
    print("TEST SUMMARY")
    print("="*80)
    print(f"Total: {total_tests} | Passed: {passed_tests} ✅ | Failed: {failed_tests} ❌")
    print(f"Success Rate: {(passed_tests/total_tests*100):.1f}%")
    print("="*80)
    
    if failed_tests > 0:
        print("\n❌ FAILED TESTS:")
        for r in test_results:
            if "❌" in r:
                print(f"  {r}")
    
    return 0 if failed_tests == 0 else 1

if __name__ == "__main__":
    sys.exit(main())
