"""Stanvard School ERP - FastAPI backend."""
import os
import logging
import hmac
import hashlib
import io
from datetime import datetime, timezone, timedelta
from typing import Optional, List, Dict, Any
from pathlib import Path

from fastapi import FastAPI, APIRouter, Depends, HTTPException, Request, Query
from fastapi.responses import StreamingResponse, JSONResponse
from starlette.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
import razorpay

from database import (
    db, mongo_client,
    schools_col, users_col, students_col, classes_col,
    fee_heads_col, fee_plans_col, fee_assignments_col,
    payments_col, receipts_col, razorpay_orders_col,
    attendance_col, homework_col, timetable_col,
    events_col, circulars_col, gallery_col, staff_col,
    notifications_col, audit_col, settings_col,
    get_next_sequence,
)
from models import (
    School, SchoolCreate, SchoolUpdate,
    User, UserCreate, UserUpdate, LoginRequest, LoginResponse,
    Student, StudentCreate, StudentUpdate,
    ClassRoom, ClassCreate,
    FeeHead, FeeHeadCreate, FeePlan, FeePlanCreate, FeeAssignment, FeeAssignmentCreate,
    FeeAssignmentUpdate, FeeAssignmentItem,
    Payment, PaymentCreate, PaymentLineItem, PaymentEdit, PaymentVoid,
    RazorpayOrderRequest, RazorpayVerifyRequest,
    AttendanceRecord, AttendanceBulkMark,
    Homework, HomeworkCreate,
    Timetable, TimetableCreate,
    Event, EventCreate,
    Circular, CircularCreate,
    GalleryAlbum, GalleryAlbumCreate,
    Staff, StaffCreate,
    Notification, NotificationCreate,
    SchoolSettings,
    now_iso, new_id,
)
from auth import (
    hash_password, verify_password, create_access_token,
    get_current_user, require_roles, current_school_id, resolve_school_id, resolve_school_id_safe,
)
from audit import log_audit
from pdf_utils import generate_receipt_pdf, generate_report_pdf

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

RAZORPAY_KEY_ID = os.environ.get('RAZORPAY_KEY_ID', '')
RAZORPAY_KEY_SECRET = os.environ.get('RAZORPAY_KEY_SECRET', '')
RAZORPAY_WEBHOOK_SECRET = os.environ.get('RAZORPAY_WEBHOOK_SECRET', '')

rzp_client = None
if RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET:
    rzp_client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))

app = FastAPI(title='Stanvard School ERP API', version='1.0.0')
api = APIRouter(prefix='/api')

logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger('stanvard')

# ----- utils -----
def strip_password(u: dict) -> dict:
    d = dict(u or {})
    d.pop('password_hash', None)
    d.pop('_id', None)
    return d


def clean(doc):
    if doc is None:
        return None
    if isinstance(doc, list):
        return [clean(d) for d in doc]
    if isinstance(doc, dict):
        return {k: v for k, v in doc.items() if k != '_id'}
    return doc


# =====================================================
# HEALTH / ROOT
# =====================================================
@api.get('/')
async def root():
    return {'app': 'Stanvard School ERP', 'status': 'running'}


@api.get('/health')
async def health():
    return {'ok': True, 'time': now_iso()}


# =====================================================
# AUTH
# =====================================================
def parent_linked_student_ids(user: Dict[str, Any]) -> List[str]:
    """Return the union of a parent user's linked student IDs.

    Merges the legacy single-child field `linked_student_id` with the
    multi-child list `linked_student_ids` so that both new and old accounts
    are supported without migration.
    """
    ids = list(user.get('linked_student_ids') or [])
    legacy = user.get('linked_student_id')
    if legacy and legacy not in ids:
        ids.append(legacy)
    return [i for i in ids if i]


def parent_can_access_student(user: Dict[str, Any], student_id: Optional[str]) -> bool:
    if not student_id:
        return False
    return student_id in parent_linked_student_ids(user)


@api.post('/auth/login', response_model=LoginResponse)
async def login(body: LoginRequest, request: Request):
    identifier = (body.email or '').strip()
    if not identifier:
        raise HTTPException(status_code=400, detail='Email or mobile number is required')

    # 1. Try email match (case-insensitive)
    user = await users_col.find_one({'email': identifier.lower()}, {'_id': 0})

    # 2. If not found and identifier looks like a phone number, try phone match.
    #    We accept +91, spaces, dashes, etc. and try both the full-digit string
    #    and the last 10 digits (Indian mobile) against the stored phone.
    if not user:
        digits_only = ''.join(ch for ch in identifier if ch.isdigit())
        candidates: List[str] = []
        if digits_only and len(digits_only) >= 7:
            candidates.append(digits_only)
            if len(digits_only) > 10:
                candidates.append(digits_only[-10:])
        for cand in candidates:
            # Exact match, then "ends-with" match on stored phone
            user = await users_col.find_one({'phone': cand}, {'_id': 0})
            if not user:
                user = await users_col.find_one(
                    {'phone': {'$regex': f'{cand}$'}}, {'_id': 0}
                )
            if user:
                break

    if not user or not verify_password(body.password, user.get('password_hash', '')):
        raise HTTPException(status_code=401, detail='Invalid credentials')
    if user.get('status') != 'active':
        raise HTTPException(status_code=403, detail='Account is inactive')
    token = create_access_token({'sub': user['id'], 'role': user['role']})
    await log_audit(action='auth.login', current_user=user,
                    ip_address=request.client.host if request.client else None)
    return LoginResponse(access_token=token, user=strip_password(user))


@api.get('/auth/me')
async def me(current=Depends(get_current_user)):
    return strip_password(current)


@api.get('/auth/my-schools')
async def my_schools(current=Depends(get_current_user)):
    """Schools accessible to current user."""
    if current['role'] == 'super_admin':
        rows = await schools_col.find({'status': {'$ne': 'deleted'}}, {'_id': 0}).to_list(1000)
    else:
        sid = current.get('school_id')
        rows = await schools_col.find({'id': sid}, {'_id': 0}).to_list(1) if sid else []
    return rows


# =====================================================
# SCHOOLS (super admin)
# =====================================================
@api.get('/schools')
async def list_schools(current=Depends(get_current_user)):
    if current['role'] == 'super_admin':
        rows = await schools_col.find({'status': {'$ne': 'deleted'}}, {'_id': 0}).to_list(1000)
    else:
        sid = current.get('school_id')
        rows = await schools_col.find({'id': sid}, {'_id': 0}).to_list(1) if sid else []
    return rows


@api.post('/schools', dependencies=[Depends(require_roles('super_admin'))])
async def create_school(body: SchoolCreate, current=Depends(get_current_user)):
    s = School(**body.model_dump())
    doc = s.model_dump()
    await schools_col.insert_one(doc)
    await log_audit(action='school.create', current_user=current,
                    entity_type='school', entity_id=s.id, details={'name': s.name})
    return clean(doc)


@api.patch('/schools/{school_id}', dependencies=[Depends(require_roles('super_admin'))])
async def update_school(school_id: str, body: SchoolUpdate, current=Depends(get_current_user)):
    upd = {k: v for k, v in body.model_dump(exclude_none=True).items()}
    upd['updated_at'] = now_iso()
    r = await schools_col.update_one({'id': school_id}, {'$set': upd})
    if not r.matched_count:
        raise HTTPException(404, 'School not found')
    await log_audit(action='school.update', current_user=current,
                    entity_type='school', entity_id=school_id, details=upd)
    doc = await schools_col.find_one({'id': school_id}, {'_id': 0})
    return clean(doc)


@api.delete('/schools/{school_id}', dependencies=[Depends(require_roles('super_admin'))])
async def archive_school(school_id: str, current=Depends(get_current_user)):
    r = await schools_col.update_one({'id': school_id},
                                     {'$set': {'status': 'archived', 'updated_at': now_iso()}})
    if not r.matched_count:
        raise HTTPException(404, 'School not found')
    await log_audit(action='school.archive', current_user=current,
                    entity_type='school', entity_id=school_id)
    return {'ok': True}


# =====================================================
# USERS
# =====================================================
@api.get('/users')
async def list_users(current=Depends(get_current_user),
                    role: Optional[str] = None,
                    school_id: Optional[str] = None):
    q: Dict[str, Any] = {}
    if role:
        q['role'] = role
    if current['role'] == 'super_admin':
        if school_id:
            q['school_id'] = school_id
    else:
        q['school_id'] = current['school_id']
    rows = await users_col.find(q, {'_id': 0}).to_list(2000)
    return [strip_password(r) for r in rows]


@api.post('/users', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_user(body: UserCreate, current=Depends(get_current_user)):
    if body.role not in {'super_admin', 'school_admin', 'accountant', 'teacher', 'parent'}:
        raise HTTPException(400, 'Invalid role')
    if current['role'] == 'school_admin':
        # school admin cannot create super_admin, and must scope to own school
        if body.role == 'super_admin':
            raise HTTPException(403, 'Cannot create super admin')
        body.school_id = current['school_id']
    exists = await users_col.find_one({'email': body.email.lower()}, {'_id': 0})
    if exists:
        raise HTTPException(400, 'Email already registered')
    u = User(
        email=body.email.lower(),
        password_hash=hash_password(body.password),
        full_name=body.full_name,
        role=body.role,
        school_id=body.school_id,
        phone=body.phone,
        linked_student_id=body.linked_student_id,
        linked_student_ids=body.linked_student_ids or [],
        linked_class_ids=body.linked_class_ids,
    )
    await users_col.insert_one(u.model_dump())
    await log_audit(action='user.create', current_user=current,
                    entity_type='user', entity_id=u.id,
                    details={'email': u.email, 'role': u.role})
    return strip_password(u.model_dump())


@api.patch('/users/{user_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def update_user(user_id: str, body: UserUpdate, current=Depends(get_current_user)):
    upd: Dict[str, Any] = {}
    for k, v in body.model_dump(exclude_none=True).items():
        if k == 'password':
            upd['password_hash'] = hash_password(v)
        else:
            upd[k] = v
    upd['updated_at'] = now_iso()
    r = await users_col.update_one({'id': user_id}, {'$set': upd})
    if not r.matched_count:
        raise HTTPException(404, 'User not found')
    await log_audit(action='user.update', current_user=current,
                    entity_type='user', entity_id=user_id, details={k: v for k, v in upd.items() if k != 'password_hash'})
    doc = await users_col.find_one({'id': user_id}, {'_id': 0})
    return strip_password(doc)


# =====================================================
# STUDENTS
# =====================================================
@api.get('/students')
async def list_students(request: Request,
                        current=Depends(get_current_user),
                        school_id: Optional[str] = None,
                        class_id: Optional[str] = None,
                        section: Optional[str] = None,
                        status: Optional[str] = None,
                        search: Optional[str] = None,
                        limit: int = 500):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid}
    # For parent role, restrict to their linked children
    if current['role'] == 'parent':
        child_ids = parent_linked_student_ids(current)
        if not child_ids:
            return []
        q['id'] = {'$in': child_ids}
    if class_id:
        q['class_id'] = class_id
    if section:
        q['section'] = section
    if status:
        q['status'] = status
    if search:
        q['$or'] = [
            {'full_name': {'$regex': search, '$options': 'i'}},
            {'admission_number': {'$regex': search, '$options': 'i'}},
            {'father_name': {'$regex': search, '$options': 'i'}},
            {'phone': {'$regex': search, '$options': 'i'}},
        ]
    rows = await students_col.find(q, {'_id': 0}).limit(limit).to_list(limit)
    return rows


@api.get('/students/{student_id}')
async def get_student(student_id: str, request: Request, current=Depends(get_current_user)):
    s = await students_col.find_one({'id': student_id}, {'_id': 0})
    if not s:
        raise HTTPException(404, 'Student not found')
    # scope check
    if current['role'] != 'super_admin' and s['school_id'] != current.get('school_id'):
        raise HTTPException(403, 'Forbidden')
    if current['role'] == 'parent' and not parent_can_access_student(current, student_id):
        raise HTTPException(403, 'Forbidden')
    return s


@api.post('/students', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_student(body: StudentCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    adm_no = body.admission_number
    if not adm_no:
        seq = await get_next_sequence(f'adm_{sid}')
        school = await schools_col.find_one({'id': sid}, {'_id': 0})
        code = (school or {}).get('code', 'STV')
        adm_no = f'{code}-2025-{seq:04d}'
    data = body.model_dump(exclude_none=True)
    data.pop('school_id', None)
    student = Student(school_id=sid, admission_number=adm_no, **{k: v for k, v in data.items() if k != 'admission_number'})
    await students_col.insert_one(student.model_dump())
    await log_audit(action='student.create', current_user=current, school_id=sid,
                    entity_type='student', entity_id=student.id,
                    details={'admission_number': adm_no, 'name': student.full_name})
    return student.model_dump()


@api.patch('/students/{student_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def update_student(student_id: str, body: StudentUpdate, current=Depends(get_current_user)):
    upd = body.model_dump(exclude_none=True)
    upd['updated_at'] = now_iso()
    r = await students_col.update_one({'id': student_id}, {'$set': upd})
    if not r.matched_count:
        raise HTTPException(404, 'Student not found')
    await log_audit(action='student.update', current_user=current,
                    entity_type='student', entity_id=student_id, details=upd)
    return await students_col.find_one({'id': student_id}, {'_id': 0})


@api.delete('/students/{student_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def delete_student(student_id: str, current=Depends(get_current_user)):
    r = await students_col.update_one({'id': student_id},
                                      {'$set': {'status': 'inactive', 'updated_at': now_iso()}})
    if not r.matched_count:
        raise HTTPException(404, 'Student not found')
    await log_audit(action='student.delete', current_user=current,
                    entity_type='student', entity_id=student_id)
    return {'ok': True}


# =====================================================
# CLASSES
# =====================================================
@api.get('/classes')
async def list_classes(request: Request, current=Depends(get_current_user),
                       school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    rows = await classes_col.find({'school_id': sid}, {'_id': 0}).to_list(1000)
    return rows


@api.post('/classes', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_class(body: ClassCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    c = ClassRoom(school_id=sid, name=body.name, sections=body.sections, teacher_id=body.teacher_id)
    await classes_col.insert_one(c.model_dump())
    await log_audit(action='class.create', current_user=current, school_id=sid,
                    entity_type='class', entity_id=c.id, details={'name': c.name})
    return c.model_dump()


# =====================================================
# FEE HEADS
# =====================================================
@api.get('/fees/heads')
async def list_fee_heads(request: Request, current=Depends(get_current_user),
                        school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    return await fee_heads_col.find({'school_id': sid}, {'_id': 0}).to_list(500)


@api.post('/fees/heads', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_fee_head(body: FeeHeadCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    fh = FeeHead(school_id=sid, name=body.name, category=body.category)
    await fee_heads_col.insert_one(fh.model_dump())
    await log_audit(action='fee_head.create', current_user=current, school_id=sid,
                    entity_type='fee_head', entity_id=fh.id, details={'name': fh.name})
    return fh.model_dump()


@api.patch('/fees/heads/{head_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def update_fee_head(head_id: str, body: FeeHeadCreate, current=Depends(get_current_user)):
    upd = body.model_dump(exclude_none=True)
    upd.pop('school_id', None)
    upd['updated_at'] = now_iso()
    r = await fee_heads_col.update_one({'id': head_id}, {'$set': upd})
    if not r.matched_count:
        raise HTTPException(404, 'Fee head not found')
    await log_audit(action='fee_head.update', current_user=current,
                    entity_type='fee_head', entity_id=head_id, details=upd)
    return await fee_heads_col.find_one({'id': head_id}, {'_id': 0})


@api.delete('/fees/heads/{head_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def delete_fee_head(head_id: str, current=Depends(get_current_user)):
    head = await fee_heads_col.find_one({'id': head_id}, {'_id': 0})
    if not head:
        raise HTTPException(404, 'Fee head not found')
    # Safety: refuse delete if this head is used in any plan or assignment.
    plan_uses = await fee_plans_col.count_documents({'items.fee_head_id': head_id})
    assign_uses = await fee_assignments_col.count_documents({'custom_items.fee_head_id': head_id})
    if plan_uses or assign_uses:
        raise HTTPException(400, f'Cannot delete: fee head is used in {plan_uses} plan(s) and {assign_uses} assignment(s). Remove references first or deactivate the head instead.')
    await fee_heads_col.delete_one({'id': head_id})
    await log_audit(action='fee_head.delete', current_user=current,
                    entity_type='fee_head', entity_id=head_id, details={'name': head.get('name')})
    return {'ok': True}


# =====================================================
# FEE PLANS
# =====================================================
@api.get('/fees/plans')
async def list_fee_plans(request: Request, current=Depends(get_current_user),
                        school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    return await fee_plans_col.find({'school_id': sid}, {'_id': 0}).to_list(500)


@api.post('/fees/plans', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_fee_plan(body: FeePlanCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    p = FeePlan(school_id=sid, name=body.name, academic_session=body.academic_session,
                class_id=body.class_id, items=body.items,
                annual_discount_percent=body.annual_discount_percent,
                late_fee_amount=body.late_fee_amount, late_fee_after_day=body.late_fee_after_day)
    await fee_plans_col.insert_one(p.model_dump())
    await log_audit(action='fee_plan.create', current_user=current, school_id=sid,
                    entity_type='fee_plan', entity_id=p.id, details={'name': p.name})
    return p.model_dump()


@api.patch('/fees/plans/{plan_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def update_fee_plan(plan_id: str, body: FeePlanCreate, current=Depends(get_current_user)):
    upd = body.model_dump(exclude_none=True)
    upd.pop('school_id', None)
    upd['updated_at'] = now_iso()
    r = await fee_plans_col.update_one({'id': plan_id}, {'$set': upd})
    if not r.matched_count:
        raise HTTPException(404, 'Plan not found')
    await log_audit(action='fee_plan.update', current_user=current,
                    entity_type='fee_plan', entity_id=plan_id, details=upd)
    return await fee_plans_col.find_one({'id': plan_id}, {'_id': 0})


@api.delete('/fees/plans/{plan_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def delete_fee_plan(plan_id: str, current=Depends(get_current_user)):
    plan = await fee_plans_col.find_one({'id': plan_id}, {'_id': 0})
    if not plan:
        raise HTTPException(404, 'Plan not found')
    used = await fee_assignments_col.count_documents({'fee_plan_id': plan_id})
    if used:
        raise HTTPException(400, f'Cannot delete: fee plan is used in {used} student assignment(s). Reassign those students first.')
    await fee_plans_col.delete_one({'id': plan_id})
    await log_audit(action='fee_plan.delete', current_user=current,
                    entity_type='fee_plan', entity_id=plan_id, details={'name': plan.get('name')})
    return {'ok': True}


@api.post('/fees/plans/{plan_id}/assign', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def assign_plan(plan_id: str, student_ids: List[str], request: Request,
                     current=Depends(get_current_user)):
    plan = await fee_plans_col.find_one({'id': plan_id}, {'_id': 0})
    if not plan:
        raise HTTPException(404, 'Plan not found')
    assignments = []
    for sid_ in student_ids:
        exists = await fee_assignments_col.find_one({'student_id': sid_, 'fee_plan_id': plan_id, 'academic_session': plan['academic_session']}, {'_id': 0})
        if exists:
            continue
        a = FeeAssignment(school_id=plan['school_id'], student_id=sid_, fee_plan_id=plan_id,
                          academic_session=plan['academic_session'])
        await fee_assignments_col.insert_one(a.model_dump())
        assignments.append(a.model_dump())
    await log_audit(action='fee_plan.assign', current_user=current, school_id=plan['school_id'],
                    entity_type='fee_plan', entity_id=plan_id,
                    details={'count': len(assignments)})
    return {'assigned': len(assignments)}


@api.get('/fees/student/{student_id}/dues')
async def student_dues(student_id: str, request: Request, current=Depends(get_current_user)):
    student = await students_col.find_one({'id': student_id}, {'_id': 0})
    if not student:
        raise HTTPException(404, 'Student not found')
    if current['role'] == 'parent' and not parent_can_access_student(current, student_id):
        raise HTTPException(403, 'Forbidden')
    assignments = await fee_assignments_col.find({'student_id': student_id}, {'_id': 0}).to_list(50)
    dues = []
    total_expected = 0.0
    total_discount = 0.0
    for a in assignments:
        # Custom items take priority over plan
        items_for_a = []
        if a.get('custom_items'):
            items_for_a = a['custom_items']
        elif a.get('fee_plan_id'):
            plan = await fee_plans_col.find_one({'id': a['fee_plan_id']}, {'_id': 0})
            if plan:
                items_for_a = plan.get('items', [])
        for item in items_for_a:
            entry = {
                'fee_head_id': item.get('fee_head_id'),
                'fee_head_name': item.get('fee_head_name'),
                'amount': item.get('amount', 0),
                'frequency': item.get('frequency', 'monthly'),
                'installments': item.get('installments', 1) if 'installments' in item else 1,
                'due_date': item.get('due_date') or a.get('due_date'),
                'assignment_id': a['id'],
                'plan_id': a.get('fee_plan_id'),
            }
            dues.append(entry)
            total_expected += entry['amount']
        # Apply assignment-level discount
        if a.get('discount_percent'):
            total_discount += total_expected * (a['discount_percent'] / 100)
        if a.get('discount_amount'):
            total_discount += a['discount_amount']
    # payments
    payments = await payments_col.find({'student_id': student_id, 'status': 'success'}, {'_id': 0}).to_list(500)
    total_paid = sum(p.get('total_paid', 0) for p in payments)
    balance = max(total_expected - total_discount - total_paid, 0)
    return {
        'student': student,
        'dues': dues,
        'assignments': assignments,
        'total_expected': total_expected,
        'total_discount': total_discount,
        'total_paid': total_paid,
        'balance': balance,
        'recent_payments': sorted(payments, key=lambda x: x.get('paid_at', ''), reverse=True)[:10],
    }


# ---------- MONTHLY / ANNUAL FEE SCHEDULE (for Parent Portal) ----------
def _session_months(session: str) -> list:
    """Return 12 (month_index, year, label) tuples for an academic session like '2026-27',
    starting from April of the first year."""
    try:
        start_year = int(session.split('-')[0])
    except Exception:
        start_year = datetime.now().year
    out = []
    for i in range(12):
        m = 4 + i  # April=4
        y = start_year + (0 if m <= 12 else 1)
        m2 = m if m <= 12 else m - 12
        y2 = start_year if m <= 12 else start_year + 1
        label = datetime(y2, m2, 1).strftime('%B %Y')
        out.append({'index': i, 'month': m2, 'year': y2, 'label': label})
    return out


def _build_month_schedule(net_annual: float, payments: list, session: str,
                          today: Optional[datetime] = None) -> tuple:
    """Compute the 12-month schedule for a student given their net annual fee
    liability and the list of *success* payments. Returns
    (schedule, monthly_amount, fully_paid, overdue_count, overdue_amount, due_amount).

    - Explicit line-item `period` labels (e.g. "April 2026") mark those months as paid.
    - Remaining paid amount is absorbed FIFO across the still-pending months.
    - Months whose (year, month) is strictly before the current month AND still
      unpaid/partial are flagged `overdue`.
    """
    monthly_amount = round(net_annual / 12.0, 2) if net_annual > 0 else 0.0
    total_paid = sum(float(p.get('total_paid') or 0) for p in payments)

    explicit_paid_labels = set()
    for p in payments:
        for it in (p.get('items') or []):
            per = (it.get('period') or '').strip()
            if per:
                explicit_paid_labels.add(per)

    months = _session_months(session)
    schedule = []
    explicit_absorbed = 0.0
    for m in months:
        entry = {
            'index': m['index'], 'label': m['label'],
            'month': m['month'], 'year': m['year'],
            'amount': monthly_amount, 'paid_amount': 0.0, 'status': 'pending',
        }
        if m['label'] in explicit_paid_labels:
            entry['status'] = 'paid'
            entry['paid_amount'] = monthly_amount
            explicit_absorbed += monthly_amount
        schedule.append(entry)

    remaining = max(total_paid - explicit_absorbed, 0.0)
    for entry in schedule:
        if entry['status'] == 'paid':
            continue
        if remaining <= 0:
            break
        if remaining >= monthly_amount - 0.01:
            entry['status'] = 'paid'
            entry['paid_amount'] = monthly_amount
            remaining -= monthly_amount
        else:
            entry['status'] = 'partial'
            entry['paid_amount'] = round(remaining, 2)
            remaining = 0

    today = today or datetime.now()
    overdue_count = 0
    overdue_amount = 0.0
    for entry in schedule:
        if entry['status'] in ('pending', 'partial'):
            if (entry['year'], entry['month']) < (today.year, today.month):
                if entry['status'] == 'pending':
                    entry['status'] = 'overdue'
                overdue_count += 1
                overdue_amount += max(monthly_amount - entry['paid_amount'], 0.0)

    fully_paid = net_annual > 0 and all(e['status'] == 'paid' for e in schedule)
    due_amount = round(max(net_annual - total_paid, 0.0), 2)
    return schedule, monthly_amount, fully_paid, overdue_count, round(overdue_amount, 2), due_amount


@api.get('/fees/student/{student_id}/fee-schedule')
async def student_fee_schedule(student_id: str, request: Request, current=Depends(get_current_user)):
    """Return a monthly & annual view of the student's fee liability for the
    parent portal. Splits the annual tuition/etc. into 12 equal monthly parts,
    marks months as paid based on prior payments (FIFO), and computes the
    'pay full' amount with the plan's annual discount applied."""
    student = await students_col.find_one({'id': student_id}, {'_id': 0})
    if not student:
        raise HTTPException(404, 'Student not found')
    if current['role'] == 'parent' and not parent_can_access_student(current, student_id):
        raise HTTPException(403, 'Forbidden')

    assignments = await fee_assignments_col.find({'student_id': student_id}, {'_id': 0}).to_list(50)

    # Aggregate: annual amount, discount, session, discount_percent from plan
    annual_total = 0.0
    plan_discount_amount = 0.0
    annual_discount_percent = 0.0
    session = '2026-27'
    items_flat: list = []  # for reference display
    for a in assignments:
        session = a.get('academic_session') or session
        # discount amount
        plan_discount_amount += float(a.get('discount_amount') or 0)
        if a.get('custom_items'):
            for it in a['custom_items']:
                items_flat.append(it)
                annual_total += float(it.get('amount') or 0)
        elif a.get('fee_plan_id'):
            plan = await fee_plans_col.find_one({'id': a['fee_plan_id']}, {'_id': 0})
            if plan:
                annual_discount_percent = max(annual_discount_percent, float(plan.get('annual_discount_percent') or 0))
                for it in (plan.get('items') or []):
                    items_flat.append(it)
                    annual_total += float(it.get('amount') or 0)
        # percent-based discount on this assignment
        if a.get('discount_percent'):
            plan_discount_amount += (annual_total * float(a['discount_percent']) / 100.0)

    # Effective yearly (net of concession)
    net_annual = max(annual_total - plan_discount_amount, 0.0)

    # Payments so far
    payments = await payments_col.find({'student_id': student_id, 'status': 'success'}, {'_id': 0}).to_list(500)
    total_paid = sum(float(p.get('total_paid') or 0) for p in payments)

    # Build 12-month schedule (shared helper)
    schedule, monthly_amount, _fully_paid, _oc, _oa, _due = _build_month_schedule(
        net_annual, payments, session
    )

    remaining_balance = max(net_annual - total_paid, 0.0)
    # Pay Full discount — apply annual_discount_percent on the REMAINING amount
    # (only meaningful before any payment; if already partially paid, apply on what's left).
    full_payment_discount = round(remaining_balance * (annual_discount_percent / 100.0), 2) if annual_discount_percent > 0 else 0.0
    payable_full = round(max(remaining_balance - full_payment_discount, 0.0), 2)

    return {
        'student': student,
        'academic_session': session,
        'annual_total': round(annual_total, 2),
        'concession': round(plan_discount_amount, 2),
        'net_annual': round(net_annual, 2),
        'monthly_amount': monthly_amount,
        'total_paid': round(total_paid, 2),
        'remaining_balance': round(remaining_balance, 2),
        'annual_discount_percent': annual_discount_percent,
        'full_payment_discount': full_payment_discount,
        'payable_full': payable_full,
        'schedule': schedule,
        'fee_head_names': list({(it.get('fee_head_name') or 'Tuition Fee') for it in items_flat}) or ['Tuition Fee'],
    }


# ---------- FEE ASSIGNMENTS (per-student) ----------
@api.get('/fees/assignments')
async def list_assignments(request: Request, current=Depends(get_current_user),
                           school_id: Optional[str] = None,
                           student_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid}
    if student_id:
        q['student_id'] = student_id
    return await fee_assignments_col.find(q, {'_id': 0}).to_list(500)


@api.post('/fees/assignments', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_assignment(body: FeeAssignmentCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    a = FeeAssignment(
        school_id=sid, student_id=body.student_id, fee_plan_id=body.fee_plan_id,
        academic_session=body.academic_session,
        custom_items=body.custom_items or [],
        custom_amount=body.custom_amount,
        discount_percent=body.discount_percent or 0.0,
        discount_amount=body.discount_amount or 0.0,
        due_date=body.due_date, remarks=body.remarks,
    )
    await fee_assignments_col.insert_one(a.model_dump())
    await log_audit(action='fee_assignment.create', current_user=current, school_id=sid,
                    entity_type='fee_assignment', entity_id=a.id,
                    details={'student_id': body.student_id, 'discount_percent': body.discount_percent,
                             'discount_amount': body.discount_amount, 'due_date': body.due_date})
    return a.model_dump()


@api.patch('/fees/assignments/{assignment_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def update_assignment(assignment_id: str, body: FeeAssignmentUpdate, current=Depends(get_current_user)):
    upd = body.model_dump(exclude_none=True)
    upd['updated_at'] = now_iso()
    r = await fee_assignments_col.update_one({'id': assignment_id}, {'$set': upd})
    if not r.matched_count:
        raise HTTPException(404, 'Assignment not found')
    await log_audit(action='fee_assignment.update', current_user=current,
                    entity_type='fee_assignment', entity_id=assignment_id, details=upd)
    return await fee_assignments_col.find_one({'id': assignment_id}, {'_id': 0})


@api.delete('/fees/assignments/{assignment_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def delete_assignment(assignment_id: str, current=Depends(get_current_user)):
    r = await fee_assignments_col.delete_one({'id': assignment_id})
    if not r.deleted_count:
        raise HTTPException(404, 'Assignment not found')
    await log_audit(action='fee_assignment.delete', current_user=current,
                    entity_type='fee_assignment', entity_id=assignment_id)
    return {'ok': True}


# =====================================================
# PAYMENTS / RECEIPTS
# =====================================================
async def _generate_receipt_number(school_id: str) -> str:
    school = await schools_col.find_one({'id': school_id}, {'_id': 0})
    code = (school or {}).get('code', 'STV')
    seq = await get_next_sequence(f'receipt_{school_id}')
    year = datetime.now().year
    return f'{code}-{year}-{seq:06d}'


async def _finalize_payment(school_id: str, body: PaymentCreate, current: dict,
                            razorpay_order_id: Optional[str] = None,
                            razorpay_payment_id: Optional[str] = None,
                            razorpay_signature: Optional[str] = None) -> Payment:
    student = await students_col.find_one({'id': body.student_id}, {'_id': 0})
    if not student:
        raise HTTPException(404, 'Student not found')
    subtotal = sum(i.amount for i in body.items)
    total_paid = subtotal + body.late_fee - body.discount
    receipt_no = await _generate_receipt_number(school_id)
    payment = Payment(
        school_id=school_id,
        student_id=student['id'],
        student_name=student['full_name'],
        receipt_number=receipt_no,
        items=body.items,
        subtotal=subtotal,
        discount=body.discount,
        late_fee=body.late_fee,
        total_paid=total_paid,
        payment_mode=body.payment_mode,
        txn_ref=body.txn_ref,
        razorpay_order_id=razorpay_order_id,
        razorpay_payment_id=razorpay_payment_id,
        razorpay_signature=razorpay_signature,
        status='success',
        remarks=body.remarks,
        collected_by_id=current.get('id'),
        collected_by_name=current.get('full_name'),
    )
    await payments_col.insert_one(payment.model_dump())
    await log_audit(action='payment.collect', current_user=current, school_id=school_id,
                    entity_type='payment', entity_id=payment.id,
                    details={'receipt_number': receipt_no, 'total': total_paid,
                             'mode': body.payment_mode, 'student': student['full_name']})
    return payment


@api.post('/payments/collect', dependencies=[Depends(require_roles('super_admin', 'school_admin', 'accountant'))])
async def collect_payment(body: PaymentCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    payment = await _finalize_payment(sid, body, current)
    return payment.model_dump()


@api.post('/payments/razorpay/order')
async def razorpay_create_order(body: RazorpayOrderRequest, request: Request,
                                current=Depends(get_current_user)):
    if not rzp_client:
        raise HTTPException(500, 'Razorpay not configured')
    student = await students_col.find_one({'id': body.student_id}, {'_id': 0})
    if not student:
        raise HTTPException(404, 'Student not found')
    if current['role'] == 'parent' and not parent_can_access_student(current, body.student_id):
        raise HTTPException(403, 'Forbidden')
    subtotal = sum(i.amount for i in body.items)
    total = subtotal + body.late_fee - body.discount
    amount_paise = int(round(total * 100))
    if amount_paise <= 0:
        raise HTTPException(400, 'Amount must be > 0')
    receipt_ref = f'stv_{new_id()[:8]}'
    order = rzp_client.order.create(data={
        'amount': amount_paise, 'currency': 'INR',
        'receipt': receipt_ref, 'payment_capture': 1,
        'notes': {
            'school_id': student['school_id'],
            'student_id': student['id'],
            'student_name': student['full_name'],
        }
    })
    await razorpay_orders_col.insert_one({
        'id': order['id'],
        'school_id': student['school_id'],
        'student_id': student['id'],
        'amount': total,
        'items': [i.model_dump() for i in body.items],
        'discount': body.discount,
        'late_fee': body.late_fee,
        'remarks': body.remarks,
        'status': 'created',
        'created_at': now_iso(),
    })
    return {
        'order_id': order['id'],
        'amount': amount_paise,
        'currency': 'INR',
        'key_id': RAZORPAY_KEY_ID,
        'student_name': student['full_name'],
    }


@api.post('/payments/razorpay/verify')
async def razorpay_verify(body: RazorpayVerifyRequest, request: Request,
                          current=Depends(get_current_user)):
    if not rzp_client:
        raise HTTPException(500, 'Razorpay not configured')
    # Verify signature
    payload = f"{body.razorpay_order_id}|{body.razorpay_payment_id}".encode()
    expected = hmac.new(RAZORPAY_KEY_SECRET.encode(), payload, hashlib.sha256).hexdigest()
    if not hmac.compare_digest(expected, body.razorpay_signature):
        raise HTTPException(400, 'Signature verification failed')
    order = await razorpay_orders_col.find_one({'id': body.razorpay_order_id}, {'_id': 0})
    if not order:
        raise HTTPException(404, 'Order not found')
    if order['status'] == 'paid':
        # idempotent
        pay = await payments_col.find_one({'razorpay_order_id': body.razorpay_order_id}, {'_id': 0})
        return pay
    pc = PaymentCreate(
        school_id=order['school_id'],
        student_id=order['student_id'],
        items=[PaymentLineItem(**i) for i in order['items']],
        discount=order.get('discount', 0),
        late_fee=order.get('late_fee', 0),
        payment_mode='razorpay',
        txn_ref=body.razorpay_payment_id,
        remarks=order.get('remarks'),
    )
    payment = await _finalize_payment(order['school_id'], pc, current,
                                      razorpay_order_id=body.razorpay_order_id,
                                      razorpay_payment_id=body.razorpay_payment_id,
                                      razorpay_signature=body.razorpay_signature)
    await razorpay_orders_col.update_one({'id': body.razorpay_order_id},
                                         {'$set': {'status': 'paid', 'payment_id': payment.id}})
    return payment.model_dump()


@api.post('/payments/razorpay/webhook')
async def razorpay_webhook(request: Request):
    body = await request.body()
    sig = request.headers.get('X-Razorpay-Signature', '')
    secret = RAZORPAY_WEBHOOK_SECRET or RAZORPAY_KEY_SECRET
    expected = hmac.new(secret.encode(), body, hashlib.sha256).hexdigest()
    if not hmac.compare_digest(expected, sig):
        raise HTTPException(400, 'Bad webhook signature')
    # Optionally handle events here
    return {'ok': True}


@api.get('/payments')
async def list_payments(request: Request, current=Depends(get_current_user),
                        school_id: Optional[str] = None,
                        student_id: Optional[str] = None,
                        start_date: Optional[str] = None,
                        end_date: Optional[str] = None,
                        mode: Optional[str] = None,
                        limit: int = 500):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid}
    if current['role'] == 'parent':
        child_ids = parent_linked_student_ids(current)
        if not child_ids:
            return []
        q['student_id'] = {'$in': child_ids}
    if student_id:
        q['student_id'] = student_id
    if mode:
        q['payment_mode'] = mode
    if start_date or end_date:
        rq: Dict[str, Any] = {}
        if start_date:
            rq['$gte'] = start_date
        if end_date:
            rq['$lte'] = end_date + 'T23:59:59'
        q['paid_at'] = rq
    rows = await payments_col.find(q, {'_id': 0}).sort('paid_at', -1).limit(limit).to_list(limit)
    return rows


@api.get('/payments/{payment_id}')
async def get_payment(payment_id: str, current=Depends(get_current_user)):
    p = await payments_col.find_one({'id': payment_id}, {'_id': 0})
    if not p:
        raise HTTPException(404, 'Payment not found')
    if current['role'] != 'super_admin' and p['school_id'] != current.get('school_id'):
        raise HTTPException(403, 'Forbidden')
    if current['role'] == 'parent' and not parent_can_access_student(current, p.get('student_id')):
        raise HTTPException(403, 'Forbidden')
    return p


@api.get('/payments/{payment_id}/receipt.pdf')
async def download_receipt(payment_id: str, current=Depends(get_current_user)):
    p = await payments_col.find_one({'id': payment_id}, {'_id': 0})
    if not p:
        raise HTTPException(404, 'Payment not found')
    if current['role'] != 'super_admin' and p['school_id'] != current.get('school_id'):
        raise HTTPException(403, 'Forbidden')
    if current['role'] == 'parent' and not parent_can_access_student(current, p.get('student_id')):
        raise HTTPException(403, 'Forbidden')
    school = await schools_col.find_one({'id': p['school_id']}, {'_id': 0}) or {}
    student = await students_col.find_one({'id': p['student_id']}, {'_id': 0}) or {}
    # enrich with class name
    if student.get('class_id'):
        cls = await classes_col.find_one({'id': student['class_id']}, {'_id': 0})
        if cls:
            student['class_name'] = cls['name']
    pdf_bytes = generate_receipt_pdf(p, school, student)
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type='application/pdf',
        headers={'Content-Disposition': f'inline; filename="receipt_{p["receipt_number"]}.pdf"'}
    )


# ---------- SUPER ADMIN: EDIT / VOID / RESTORE RECEIPTS ----------
def _payment_snapshot(p: dict) -> dict:
    """Return an immutable snapshot of the mutable financial fields of a payment.
    Stored inside `edit_history` for a full audit trail of every super-admin edit."""
    return {
        'items': p.get('items') or [],
        'subtotal': p.get('subtotal'),
        'discount': p.get('discount'),
        'late_fee': p.get('late_fee'),
        'total_paid': p.get('total_paid'),
        'payment_mode': p.get('payment_mode'),
        'txn_ref': p.get('txn_ref'),
        'remarks': p.get('remarks'),
        'paid_at': p.get('paid_at'),
        'status': p.get('status'),
    }


@api.patch('/payments/{payment_id}',
           dependencies=[Depends(require_roles('super_admin'))])
async def edit_payment(payment_id: str, body: PaymentEdit,
                       current=Depends(get_current_user)):
    """Super-admin only. Edit a previously collected payment (fix wrong items,
    amounts, mode, etc.). The `receipt_number` is preserved for financial
    continuity; every change is recorded in `edit_history` for audit."""
    p = await payments_col.find_one({'id': payment_id}, {'_id': 0})
    if not p:
        raise HTTPException(404, 'Payment not found')
    if not (body.reason and body.reason.strip()):
        raise HTTPException(400, 'Reason is required for editing a receipt')
    if p.get('status') == 'voided':
        raise HTTPException(400, 'Cannot edit a voided receipt. Restore it first.')

    # Snapshot BEFORE mutation.
    prev = _payment_snapshot(p)
    prev['edited_at'] = now_iso()
    prev['edited_by_id'] = current.get('id')
    prev['edited_by_name'] = current.get('full_name')
    prev['reason'] = body.reason.strip()

    update: Dict[str, Any] = {}
    if body.items is not None:
        items_dump = [i.model_dump() for i in body.items]
        subtotal = sum(float(i.get('amount') or 0) for i in items_dump)
        discount = float(body.discount if body.discount is not None else (p.get('discount') or 0))
        late_fee = float(body.late_fee if body.late_fee is not None else (p.get('late_fee') or 0))
        update['items'] = items_dump
        update['subtotal'] = subtotal
        update['discount'] = discount
        update['late_fee'] = late_fee
        update['total_paid'] = round(subtotal + late_fee - discount, 2)
    else:
        # items unchanged, but discount / late_fee may still change → recompute total
        if body.discount is not None or body.late_fee is not None:
            subtotal = float(p.get('subtotal') or 0)
            discount = float(body.discount if body.discount is not None else (p.get('discount') or 0))
            late_fee = float(body.late_fee if body.late_fee is not None else (p.get('late_fee') or 0))
            update['discount'] = discount
            update['late_fee'] = late_fee
            update['total_paid'] = round(subtotal + late_fee - discount, 2)

    if body.payment_mode is not None:
        update['payment_mode'] = body.payment_mode
    if body.txn_ref is not None:
        update['txn_ref'] = body.txn_ref
    if body.remarks is not None:
        update['remarks'] = body.remarks
    if body.paid_at is not None and body.paid_at.strip():
        update['paid_at'] = body.paid_at

    update['edited_at'] = prev['edited_at']
    update['edited_by_id'] = current.get('id')
    update['edited_by_name'] = current.get('full_name')
    update['edited_reason'] = body.reason.strip()

    await payments_col.update_one(
        {'id': payment_id},
        {'$set': update, '$push': {'edit_history': prev}},
    )

    await log_audit(action='payment.edit', current_user=current,
                    school_id=p.get('school_id'),
                    entity_type='payment', entity_id=payment_id,
                    details={
                        'receipt_number': p.get('receipt_number'),
                        'reason': body.reason.strip(),
                        'previous_total': prev.get('total_paid'),
                        'new_total': update.get('total_paid', prev.get('total_paid')),
                    })
    return await payments_col.find_one({'id': payment_id}, {'_id': 0})


@api.post('/payments/{payment_id}/void',
          dependencies=[Depends(require_roles('super_admin'))])
async def void_payment(payment_id: str, body: PaymentVoid,
                       current=Depends(get_current_user)):
    """Super-admin only. Void / cancel a receipt. The payment is retained for
    audit but its `status` is set to `voided`, which automatically excludes it
    from all fee-schedule / dues / collection-report aggregations."""
    p = await payments_col.find_one({'id': payment_id}, {'_id': 0})
    if not p:
        raise HTTPException(404, 'Payment not found')
    if p.get('status') == 'voided':
        raise HTTPException(400, 'Receipt is already voided')
    if not (body.reason and body.reason.strip()):
        raise HTTPException(400, 'Reason is required to void a receipt')

    ts = now_iso()
    await payments_col.update_one(
        {'id': payment_id},
        {'$set': {
            'status': 'voided',
            'voided_at': ts,
            'voided_by_id': current.get('id'),
            'voided_by_name': current.get('full_name'),
            'void_reason': body.reason.strip(),
        }},
    )
    await log_audit(action='payment.void', current_user=current,
                    school_id=p.get('school_id'),
                    entity_type='payment', entity_id=payment_id,
                    details={
                        'receipt_number': p.get('receipt_number'),
                        'reason': body.reason.strip(),
                        'amount_reversed': p.get('total_paid'),
                        'student_id': p.get('student_id'),
                    })
    return await payments_col.find_one({'id': payment_id}, {'_id': 0})


@api.post('/payments/{payment_id}/restore',
          dependencies=[Depends(require_roles('super_admin'))])
async def restore_payment(payment_id: str, current=Depends(get_current_user)):
    """Super-admin only. Un-void a previously voided receipt (mistake reversal)."""
    p = await payments_col.find_one({'id': payment_id}, {'_id': 0})
    if not p:
        raise HTTPException(404, 'Payment not found')
    if p.get('status') != 'voided':
        raise HTTPException(400, 'Receipt is not voided')

    await payments_col.update_one(
        {'id': payment_id},
        {'$set': {'status': 'success'},
         '$unset': {'voided_at': '', 'voided_by_id': '',
                    'voided_by_name': '', 'void_reason': ''}},
    )
    await log_audit(action='payment.restore', current_user=current,
                    school_id=p.get('school_id'),
                    entity_type='payment', entity_id=payment_id,
                    details={'receipt_number': p.get('receipt_number')})
    return await payments_col.find_one({'id': payment_id}, {'_id': 0})


# =====================================================
# ATTENDANCE
# =====================================================
@api.post('/attendance/mark', dependencies=[Depends(require_roles('super_admin', 'school_admin', 'teacher'))])
async def mark_attendance(body: AttendanceBulkMark, request: Request,
                          current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    # Delete existing for that date/class first (upsert semantics)
    await attendance_col.delete_many({
        'school_id': sid,
        'date': body.date,
        'class_id': body.class_id,
        'section': body.section,
    })
    docs = []
    for e in body.entries:
        rec = AttendanceRecord(
            school_id=sid, date=body.date, class_id=body.class_id,
            section=body.section, student_id=e.get('student_id'),
            status=e.get('status', 'present'),
            remarks=e.get('remarks'),
            marked_by_id=current.get('id'),
        )
        docs.append(rec.model_dump())
    if docs:
        await attendance_col.insert_many(docs)
    await log_audit(action='attendance.mark', current_user=current, school_id=sid,
                    entity_type='attendance', entity_id=f"{body.date}-{body.class_id}",
                    details={'count': len(docs), 'date': body.date})
    return {'saved': len(docs)}


@api.get('/attendance')
async def list_attendance(request: Request, current=Depends(get_current_user),
                          school_id: Optional[str] = None,
                          date: Optional[str] = None,
                          class_id: Optional[str] = None,
                          section: Optional[str] = None,
                          student_id: Optional[str] = None,
                          start_date: Optional[str] = None,
                          end_date: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid}
    if date:
        q['date'] = date
    if class_id:
        q['class_id'] = class_id
    if section:
        q['section'] = section
    if student_id:
        q['student_id'] = student_id
    if current['role'] == 'parent':
        child_ids = parent_linked_student_ids(current)
        if not child_ids:
            return []
        q['student_id'] = {'$in': child_ids}
    if start_date or end_date:
        rq: Dict[str, Any] = {}
        if start_date:
            rq['$gte'] = start_date
        if end_date:
            rq['$lte'] = end_date
        q['date'] = rq if rq else q.get('date')
    rows = await attendance_col.find(q, {'_id': 0}).limit(5000).to_list(5000)
    return rows


# =====================================================
# HOMEWORK
# =====================================================
@api.get('/homework')
async def list_homework(request: Request, current=Depends(get_current_user),
                        school_id: Optional[str] = None,
                        class_id: Optional[str] = None,
                        section: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid}
    if class_id:
        q['class_id'] = class_id
    if section:
        q['section'] = section
    if current['role'] == 'parent':
        child_ids = parent_linked_student_ids(current)
        if not child_ids:
            return []
        # Homework is class-scoped; pull the child's classes and match
        children = await students_col.find(
            {'id': {'$in': child_ids}}, {'_id': 0, 'class_id': 1, 'section': 1}
        ).to_list(100)
        class_ids = list({c.get('class_id') for c in children if c.get('class_id')})
        if not class_ids:
            return []
        q['class_id'] = {'$in': class_ids}
    rows = await homework_col.find(q, {'_id': 0}).sort('created_at', -1).limit(500).to_list(500)
    return rows


@api.post('/homework', dependencies=[Depends(require_roles('super_admin', 'school_admin', 'teacher'))])
async def create_homework(body: HomeworkCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    hw = Homework(school_id=sid, class_id=body.class_id, section=body.section,
                  subject=body.subject, title=body.title, description=body.description,
                  due_date=body.due_date, attachment_url=body.attachment_url,
                  created_by_id=current.get('id'),
                  created_by_name=current.get('full_name'))
    await homework_col.insert_one(hw.model_dump())
    await log_audit(action='homework.create', current_user=current, school_id=sid,
                    entity_type='homework', entity_id=hw.id, details={'title': hw.title})
    return hw.model_dump()


# =====================================================
# TIMETABLE
# =====================================================
@api.get('/timetable')
async def get_timetable(request: Request, current=Depends(get_current_user),
                        school_id: Optional[str] = None,
                        class_id: Optional[str] = None,
                        section: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid}
    if class_id:
        q['class_id'] = class_id
    if section:
        q['section'] = section
    return await timetable_col.find(q, {'_id': 0}).to_list(200)


@api.post('/timetable', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def upsert_timetable(body: TimetableCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    exists = await timetable_col.find_one({'school_id': sid, 'class_id': body.class_id, 'section': body.section}, {'_id': 0})
    slots = [s.model_dump() for s in body.slots]
    if exists:
        await timetable_col.update_one({'id': exists['id']}, {'$set': {'slots': slots, 'updated_at': now_iso()}})
        return await timetable_col.find_one({'id': exists['id']}, {'_id': 0})
    t = Timetable(school_id=sid, class_id=body.class_id, section=body.section, slots=body.slots)
    await timetable_col.insert_one(t.model_dump())
    return t.model_dump()


# =====================================================
# EVENTS, CIRCULARS, GALLERY, STAFF, NOTIFICATIONS
# =====================================================
@api.get('/events')
async def list_events(request: Request, current=Depends(get_current_user), school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    return await events_col.find({'school_id': sid}, {'_id': 0}).sort('event_date', -1).limit(200).to_list(200)


@api.post('/events', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_event(body: EventCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    e = Event(school_id=sid, title=body.title, description=body.description,
              event_date=body.event_date, location=body.location, image_url=body.image_url)
    await events_col.insert_one(e.model_dump())
    await log_audit(action='event.create', current_user=current, school_id=sid,
                    entity_type='event', entity_id=e.id, details={'title': e.title})
    return e.model_dump()


@api.get('/circulars')
async def list_circulars(request: Request, current=Depends(get_current_user), school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    return await circulars_col.find({'school_id': sid, 'status': 'published'}, {'_id': 0}).sort('created_at', -1).limit(200).to_list(200)


@api.post('/circulars', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_circular(body: CircularCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    c = Circular(school_id=sid, title=body.title, body=body.body, priority=body.priority,
                 status=body.status, publish_at=body.publish_at,
                 attachment_url=body.attachment_url, audience=body.audience,
                 class_id=body.class_id, created_by_name=current.get('full_name'))
    await circulars_col.insert_one(c.model_dump())
    await log_audit(action='circular.create', current_user=current, school_id=sid,
                    entity_type='circular', entity_id=c.id, details={'title': c.title})
    return c.model_dump()


@api.get('/gallery')
async def list_gallery(request: Request, current=Depends(get_current_user), school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    return await gallery_col.find({'school_id': sid}, {'_id': 0}).sort('created_at', -1).to_list(200)


@api.post('/gallery', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_album(body: GalleryAlbumCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    a = GalleryAlbum(school_id=sid, title=body.title, description=body.description,
                     cover_url=body.cover_url, photos=body.photos)
    await gallery_col.insert_one(a.model_dump())
    await log_audit(action='gallery.create', current_user=current, school_id=sid,
                    entity_type='gallery', entity_id=a.id, details={'title': a.title})
    return a.model_dump()


@api.get('/staff')
async def list_staff(request: Request, current=Depends(get_current_user), school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    return await staff_col.find({'school_id': sid}, {'_id': 0}).to_list(500)


@api.post('/staff', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def create_staff(body: StaffCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    s = Staff(school_id=sid, full_name=body.full_name, email=body.email, phone=body.phone,
              designation=body.designation, department=body.department,
              subjects=body.subjects, joining_date=body.joining_date, photo_url=body.photo_url)
    await staff_col.insert_one(s.model_dump())
    await log_audit(action='staff.create', current_user=current, school_id=sid,
                    entity_type='staff', entity_id=s.id, details={'name': s.full_name})
    return s.model_dump()


@api.get('/notifications')
async def list_notifications(request: Request, current=Depends(get_current_user), school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid}
    # audience filter for role-scoped users
    if current['role'] == 'parent':
        child_ids = parent_linked_student_ids(current)
        q['$or'] = [{'audience': 'all'}, {'audience': 'parents'},
                    {'student_ids': {'$in': child_ids}} if child_ids else {'student_ids': None}]
    elif current['role'] == 'teacher':
        q['$or'] = [{'audience': 'all'}, {'audience': 'teachers'}]
    return await notifications_col.find(q, {'_id': 0}).sort('created_at', -1).limit(200).to_list(200)


@api.post('/notifications', dependencies=[Depends(require_roles('super_admin', 'school_admin', 'teacher'))])
async def create_notification(body: NotificationCreate, request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, body.school_id, request.headers.get('X-School-Id'))
    n = Notification(school_id=sid, title=body.title, body=body.body,
                     audience=body.audience, class_id=body.class_id,
                     student_ids=body.student_ids, kind=body.kind)
    await notifications_col.insert_one(n.model_dump())
    await log_audit(action='notification.create', current_user=current, school_id=sid,
                    entity_type='notification', entity_id=n.id,
                    details={'title': n.title, 'audience': n.audience})
    return n.model_dump()


# =====================================================
# REPORTS
# =====================================================
@api.get('/reports/collection')
async def report_collection(request: Request, current=Depends(get_current_user),
                            school_id: Optional[str] = None,
                            start_date: Optional[str] = None,
                            end_date: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid, 'status': 'success'}
    if start_date or end_date:
        rq: Dict[str, Any] = {}
        if start_date:
            rq['$gte'] = start_date
        if end_date:
            rq['$lte'] = end_date + 'T23:59:59'
        q['paid_at'] = rq
    payments = await payments_col.find(q, {'_id': 0}).sort('paid_at', -1).to_list(5000)
    total = sum(p.get('total_paid', 0) for p in payments)
    by_mode: Dict[str, float] = {}
    for p in payments:
        m = p.get('payment_mode', 'cash')
        by_mode[m] = by_mode.get(m, 0) + p.get('total_paid', 0)
    return {'payments': payments, 'total': total, 'by_mode': by_mode, 'count': len(payments)}


@api.get('/reports/collection.pdf')
async def report_collection_pdf(request: Request, current=Depends(get_current_user),
                                school_id: Optional[str] = None,
                                start_date: Optional[str] = None,
                                end_date: Optional[str] = None):
    data = await report_collection(request, current, school_id, start_date, end_date)
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    school = await schools_col.find_one({'id': sid}, {'_id': 0}) or {}
    cols = ['Receipt No', 'Date', 'Student', 'Mode', 'Amount (Rs.)']
    rows = []
    for p in data['payments']:
        rows.append([p.get('receipt_number', ''),
                     (p.get('paid_at') or '')[:10],
                     p.get('student_name', ''),
                     str(p.get('payment_mode', '')).replace('_', ' ').title(),
                     f"{p.get('total_paid', 0):,.2f}"])
    subtitle = f"Period: {start_date or 'All'}  to  {end_date or 'Today'}"
    summary = {'Total Collection': f"Rs. {data['total']:,.2f}", 'Transactions': data['count']}
    pdf = generate_report_pdf('Fee Collection Report', subtitle, cols, rows,
                              school.get('name', 'Stanvard School'), summary)
    return StreamingResponse(io.BytesIO(pdf), media_type='application/pdf',
                             headers={'Content-Disposition': 'inline; filename="collection_report.pdf"'})


@api.get('/reports/collection.csv')
async def report_collection_csv(request: Request, current=Depends(get_current_user),
                                school_id: Optional[str] = None,
                                start_date: Optional[str] = None,
                                end_date: Optional[str] = None):
    data = await report_collection(request, current, school_id, start_date, end_date)
    import csv
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['Receipt No', 'Date', 'Student', 'Mode', 'Amount'])
    for p in data['payments']:
        writer.writerow([p.get('receipt_number', ''),
                        (p.get('paid_at') or '')[:10],
                        p.get('student_name', ''),
                        p.get('payment_mode', ''),
                        p.get('total_paid', 0)])
    return StreamingResponse(io.BytesIO(buffer.getvalue().encode()),
                             media_type='text/csv',
                             headers={'Content-Disposition': 'attachment; filename="collection.csv"'})


@api.get('/reports/collection.xlsx')
async def report_collection_xlsx(request: Request, current=Depends(get_current_user),
                                 school_id: Optional[str] = None,
                                 start_date: Optional[str] = None,
                                 end_date: Optional[str] = None):
    from openpyxl import Workbook
    data = await report_collection(request, current, school_id, start_date, end_date)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Collection'
    ws.append(['Receipt No', 'Date', 'Student', 'Mode', 'Amount'])
    for p in data['payments']:
        ws.append([p.get('receipt_number', ''),
                   (p.get('paid_at') or '')[:10],
                   p.get('student_name', ''),
                   p.get('payment_mode', ''),
                   p.get('total_paid', 0)])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             headers={'Content-Disposition': 'attachment; filename="collection.xlsx"'})


@api.get('/reports/attendance')
async def report_attendance(request: Request, current=Depends(get_current_user),
                            school_id: Optional[str] = None,
                            start_date: Optional[str] = None,
                            end_date: Optional[str] = None,
                            class_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    q: Dict[str, Any] = {'school_id': sid}
    if start_date or end_date:
        rq: Dict[str, Any] = {}
        if start_date:
            rq['$gte'] = start_date
        if end_date:
            rq['$lte'] = end_date
        q['date'] = rq
    if class_id:
        q['class_id'] = class_id
    rows = await attendance_col.find(q, {'_id': 0}).to_list(20000)
    total = len(rows)
    present = sum(1 for r in rows if r.get('status') == 'present')
    absent = sum(1 for r in rows if r.get('status') == 'absent')
    leave = sum(1 for r in rows if r.get('status') == 'leave')
    return {'total': total, 'present': present, 'absent': absent, 'leave': leave,
            'attendance_rate': round(present / total * 100, 2) if total else 0.0}


# =====================================================
# DASHBOARD
# =====================================================
@api.get('/dashboard/summary')
async def dashboard_summary(request: Request, current=Depends(get_current_user),
                            school_id: Optional[str] = None):
    sid = await resolve_school_id_safe(current, school_id, request.headers.get('X-School-Id'))
    today = datetime.now().strftime('%Y-%m-%d')
    month_start = datetime.now().strftime('%Y-%m-01')

    total_students = await students_col.count_documents({'school_id': sid, 'status': 'active'})
    total_staff = await staff_col.count_documents({'school_id': sid, 'status': 'active'})

    # Today's collection (exclude voided receipts)
    today_payments = await payments_col.find({'school_id': sid, 'status': 'success', 'paid_at': {'$gte': today, '$lt': today + 'T23:59:59'}}, {'_id': 0}).to_list(2000)
    today_collection = sum(p.get('total_paid', 0) for p in today_payments)
    # Monthly collection (exclude voided receipts)
    month_payments = await payments_col.find({'school_id': sid, 'status': 'success', 'paid_at': {'$gte': month_start}}, {'_id': 0}).to_list(5000)
    monthly_collection = sum(p.get('total_paid', 0) for p in month_payments)

    # Attendance today
    today_att = await attendance_col.find({'school_id': sid, 'date': today}, {'_id': 0}).to_list(5000)
    present_today = sum(1 for a in today_att if a.get('status') == 'present')
    absent_today = sum(1 for a in today_att if a.get('status') == 'absent')

    # New admissions (this month)
    new_admissions = await students_col.count_documents({
        'school_id': sid,
        'created_at': {'$gte': month_start}
    })

    # Recent activity
    recent_payments = await payments_col.find({'school_id': sid}, {'_id': 0}).sort('paid_at', -1).limit(5).to_list(5)
    upcoming_events = await events_col.find({
        'school_id': sid, 'event_date': {'$gte': today}
    }, {'_id': 0}).sort('event_date', 1).limit(5).to_list(5)
    recent_circulars = await circulars_col.find({
        'school_id': sid, 'status': 'published'
    }, {'_id': 0}).sort('created_at', -1).limit(5).to_list(5)

    # Pending fees (naive: students with assignments who haven't paid this month) — exclude voided
    with_assignments = await fee_assignments_col.distinct('student_id', {'school_id': sid})
    paid_this_month = await payments_col.distinct('student_id', {'school_id': sid, 'status': 'success', 'paid_at': {'$gte': month_start}})
    pending_students = max(len(with_assignments) - len(paid_this_month), 0)

    # Collection trend last 7 days (exclude voided)
    trend = []
    for i in range(6, -1, -1):
        d = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
        payments = await payments_col.find({'school_id': sid, 'status': 'success', 'paid_at': {'$gte': d, '$lt': d + 'T23:59:59'}}, {'_id': 0}).to_list(1000)
        trend.append({'date': d, 'amount': sum(p.get('total_paid', 0) for p in payments)})

    return {
        'total_students': total_students,
        'total_staff': total_staff,
        'today_collection': today_collection,
        'monthly_collection': monthly_collection,
        'present_today': present_today,
        'absent_today': absent_today,
        'new_admissions': new_admissions,
        'pending_students': pending_students,
        'recent_payments': recent_payments,
        'upcoming_events': upcoming_events,
        'recent_circulars': recent_circulars,
        'collection_trend': trend,
    }


# =====================================================
# AUDIT LOGS
# =====================================================
@api.get('/audit-logs', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def list_audit(request: Request, current=Depends(get_current_user),
                    school_id: Optional[str] = None, limit: int = 200):
    q: Dict[str, Any] = {}
    if current['role'] == 'super_admin':
        if school_id:
            q['school_id'] = school_id
    else:
        q['school_id'] = current['school_id']
    rows = await audit_col.find(q, {'_id': 0}).sort('created_at', -1).limit(limit).to_list(limit)
    return rows


# =====================================================
# SETTINGS
# =====================================================
@api.get('/settings')
async def get_settings(request: Request, current=Depends(get_current_user), school_id: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    s = await settings_col.find_one({'school_id': sid}, {'_id': 0})
    if not s:
        s = SchoolSettings(school_id=sid).model_dump()
        await settings_col.insert_one(dict(s))
        s = await settings_col.find_one({'school_id': sid}, {'_id': 0})
    return s


@api.patch('/settings', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def update_settings(payload: Dict[str, Any], request: Request, current=Depends(get_current_user)):
    sid = resolve_school_id(current, None, request.headers.get('X-School-Id'))
    payload['updated_at'] = now_iso()
    await settings_col.update_one({'school_id': sid}, {'$set': payload}, upsert=True)
    await log_audit(action='settings.update', current_user=current, school_id=sid,
                    entity_type='settings', details=payload)
    return await settings_col.find_one({'school_id': sid}, {'_id': 0})


# =====================================================
# USER DELETE + PASSWORD RESET
# =====================================================
@api.delete('/users/{user_id}', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def delete_user(user_id: str, current=Depends(get_current_user)):
    target = await users_col.find_one({'id': user_id}, {'_id': 0})
    if not target:
        raise HTTPException(404, 'User not found')
    if current['role'] == 'school_admin':
        if target.get('school_id') != current.get('school_id') or target.get('role') == 'super_admin':
            raise HTTPException(403, 'Cannot delete this user')
    if target['id'] == current['id']:
        raise HTTPException(400, 'Cannot delete yourself')
    await users_col.update_one({'id': user_id}, {'$set': {'status': 'inactive', 'updated_at': now_iso()}})
    await log_audit(action='user.delete', current_user=current,
                    entity_type='user', entity_id=user_id, details={'email': target.get('email')})
    return {'ok': True}


@api.post('/users/{user_id}/reset-password', dependencies=[Depends(require_roles('super_admin', 'school_admin'))])
async def reset_password(user_id: str, payload: Dict[str, str], current=Depends(get_current_user)):
    new_pw = payload.get('password')
    if not new_pw or len(new_pw) < 6:
        raise HTTPException(400, 'Password must be at least 6 characters')
    target = await users_col.find_one({'id': user_id}, {'_id': 0})
    if not target:
        raise HTTPException(404, 'User not found')
    if current['role'] == 'school_admin' and target.get('school_id') != current.get('school_id'):
        raise HTTPException(403, 'Forbidden')
    await users_col.update_one({'id': user_id}, {'$set': {'password_hash': hash_password(new_pw), 'updated_at': now_iso()}})
    await log_audit(action='user.reset_password', current_user=current,
                    entity_type='user', entity_id=user_id, details={'email': target.get('email')})
    return {'ok': True}


# =====================================================
# ANALYTICS
# =====================================================
@api.get('/analytics')
async def analytics(request: Request, current=Depends(get_current_user),
                   school_id: Optional[str] = None, year: Optional[int] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    year = year or datetime.now().year
    year_start = f'{year}-01-01'
    year_end = f'{year}-12-31T23:59:59'

    # All payments for the year
    payments = await payments_col.find({
        'school_id': sid, 'status': 'success',
        'paid_at': {'$gte': year_start, '$lte': year_end}
    }, {'_id': 0}).to_list(20000)

    # Monthly breakdown (Jan..Dec)
    months = [{'month': datetime(year, m, 1).strftime('%b'),
               'received': 0.0, 'transactions': 0, 'discount': 0.0, 'late_fee': 0.0}
              for m in range(1, 13)]
    by_mode = {}
    by_head = {}
    for p in payments:
        try:
            paid_month = int((p.get('paid_at') or '')[5:7])
            months[paid_month - 1]['received'] += p.get('total_paid', 0)
            months[paid_month - 1]['transactions'] += 1
            months[paid_month - 1]['discount'] += p.get('discount', 0)
            months[paid_month - 1]['late_fee'] += p.get('late_fee', 0)
        except Exception:
            pass
        mode = p.get('payment_mode', 'cash')
        by_mode[mode] = by_mode.get(mode, 0) + p.get('total_paid', 0)
        for item in p.get('items', []):
            hd = item.get('fee_head_name', 'Other')
            by_head[hd] = by_head.get(hd, 0) + item.get('amount', 0)

    total_received = sum(p.get('total_paid', 0) for p in payments)
    total_transactions = len(payments)
    total_discount = sum(p.get('discount', 0) for p in payments)
    total_late_fee = sum(p.get('late_fee', 0) for p in payments)

    # Expected: sum of all assignment amounts (custom_items or plan)
    assignments = await fee_assignments_col.find({'school_id': sid}, {'_id': 0}).to_list(10000)
    total_expected = 0.0
    plans_cache: Dict[str, Any] = {}
    for a in assignments:
        items = a.get('custom_items') or []
        if not items and a.get('fee_plan_id'):
            pid = a['fee_plan_id']
            if pid not in plans_cache:
                plans_cache[pid] = await fee_plans_col.find_one({'id': pid}, {'_id': 0}) or {}
            items = plans_cache[pid].get('items', [])
        for it in items:
            total_expected += it.get('amount', 0)
        total_expected -= a.get('discount_amount', 0)
    total_due = max(total_expected - total_received, 0)

    # Attendance summary for the year
    att = await attendance_col.find({'school_id': sid, 'date': {'$gte': year_start[:10], '$lte': year_end[:10]}}, {'_id': 0}).to_list(50000)
    att_total = len(att)
    att_present = sum(1 for a in att if a.get('status') == 'present')
    att_absent = sum(1 for a in att if a.get('status') == 'absent')
    att_leave = sum(1 for a in att if a.get('status') == 'leave')

    # New admissions per month
    students = await students_col.find({'school_id': sid}, {'_id': 0}).to_list(5000)
    adm_months = [0] * 12
    for s in students:
        try:
            ca = s.get('created_at') or s.get('admission_date') or ''
            if ca and int(ca[:4]) == year:
                adm_months[int(ca[5:7]) - 1] += 1
        except Exception:
            pass
    for i, m in enumerate(months):
        m['admissions'] = adm_months[i]

    # Class-wise pending (approximation)
    classes = await classes_col.find({'school_id': sid}, {'_id': 0}).to_list(200)
    by_class = []
    for c in classes:
        students_in_class = [s for s in students if s.get('class_id') == c['id']]
        payer_ids = {p['student_id'] for p in payments if p['student_id'] in {s['id'] for s in students_in_class}}
        by_class.append({
            'class_id': c['id'], 'class_name': c['name'],
            'total_students': len(students_in_class),
            'paying_students': len(payer_ids),
            'pending_students': max(len(students_in_class) - len(payer_ids), 0),
        })

    return {
        'year': year,
        'total_received': total_received,
        'total_expected': total_expected,
        'total_due': total_due,
        'total_discount': total_discount,
        'total_late_fee': total_late_fee,
        'total_transactions': total_transactions,
        'months': months,
        'by_mode': by_mode,
        'by_head': by_head,
        'by_class': by_class,
        'attendance': {
            'total': att_total,
            'present': att_present,
            'absent': att_absent,
            'leave': att_leave,
            'rate': round(att_present / att_total * 100, 2) if att_total else 0.0,
        },
        'total_students': len(students),
    }


# =====================================================
# ENHANCED REPORTS - Fee Status per student (paid/pending)
# =====================================================
@api.get('/reports/fee-status', dependencies=[Depends(require_roles('super_admin', 'school_admin', 'accountant'))])
async def report_fee_status(request: Request, current=Depends(get_current_user),
                            school_id: Optional[str] = None,
                            class_id: Optional[str] = None,
                            section: Optional[str] = None,
                            class_sections: Optional[str] = None,
                            status_filter: Optional[str] = None,   # paid|partial|unpaid
                            due_min: Optional[float] = None,
                            due_max: Optional[float] = None,
                            payment_date_start: Optional[str] = None,
                            payment_date_end: Optional[str] = None,
                            quick_view: Optional[str] = None,      # defaulters|fully_paid|upcoming
                            behavior: Optional[str] = None):        # regular|late|defaulter
    """Student fee status report.

    Supports both:
      - Legacy single-filter: `class_id` + `section` (kept for backwards compatibility)
      - New multi-select: `class_sections` = comma-separated pairs like
        "<class_id>:<section>,<class_id>:,<class_id>:A"
        A blank section (e.g. "abc:") means "all sections of that class".

    Extended fields per row:
      - last_payment_date (ISO date str or None)
      - overdue_days (int, 0 if not overdue)
      - behavior_tag ('regular' | 'late' | 'defaulter' | 'na')
      - upcoming_due_date (nearest future due date, if any)
      - collection_percent
    Plus class/section rollups in `by_class`.
    """
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))

    cs_pairs: List[Dict[str, Optional[str]]] = []
    if class_sections:
        for token in class_sections.split(','):
            token = token.strip()
            if not token:
                continue
            if ':' in token:
                cid, sec = token.split(':', 1)
                cs_pairs.append({'class_id': cid.strip(), 'section': (sec.strip() or None)})
            else:
                cs_pairs.append({'class_id': token, 'section': None})

    student_q: Dict[str, Any] = {'school_id': sid, 'status': 'active'}
    if cs_pairs:
        or_clauses: List[Dict[str, Any]] = []
        for p in cs_pairs:
            clause: Dict[str, Any] = {'class_id': p['class_id']}
            if p['section']:
                clause['section'] = p['section']
            or_clauses.append(clause)
        student_q['$or'] = or_clauses
    else:
        if class_id:
            student_q['class_id'] = class_id
        if section:
            student_q['section'] = section
    students = await students_col.find(student_q, {'_id': 0}).to_list(5000)

    classes = await classes_col.find({'school_id': sid}, {'_id': 0}).to_list(200)
    class_map = {c['id']: c['name'] for c in classes}

    student_ids = [s['id'] for s in students]
    assignments = await fee_assignments_col.find({'student_id': {'$in': student_ids}}, {'_id': 0}).to_list(20000)
    payments = await payments_col.find({'student_id': {'$in': student_ids}, 'status': 'success'}, {'_id': 0}).to_list(20000)

    plan_cache: Dict[str, Any] = {}
    async def _get_plan(pid):
        if pid not in plan_cache:
            plan_cache[pid] = await fee_plans_col.find_one({'id': pid}, {'_id': 0}) or {}
        return plan_cache[pid]

    expected_by_stu: Dict[str, float] = {}
    discount_by_stu: Dict[str, float] = {}
    due_date_by_stu: Dict[str, Optional[str]] = {}
    for a in assignments:
        sid_ = a['student_id']
        items = a.get('custom_items') or []
        if not items and a.get('fee_plan_id'):
            plan = await _get_plan(a['fee_plan_id'])
            items = plan.get('items', [])
        expected_by_stu[sid_] = expected_by_stu.get(sid_, 0.0) + sum(i.get('amount', 0) for i in items)
        discount_by_stu[sid_] = discount_by_stu.get(sid_, 0.0) + a.get('discount_amount', 0)
        if a.get('due_date') and not due_date_by_stu.get(sid_):
            due_date_by_stu[sid_] = a['due_date']

    # Payment aggregates: total paid, last date, late flag, per-mode counts
    paid_by_stu: Dict[str, float] = {}
    last_paid_by_stu: Dict[str, Optional[str]] = {}
    late_flag_by_stu: Dict[str, bool] = {}
    for p in payments:
        sid_ = p['student_id']
        paid_by_stu[sid_] = paid_by_stu.get(sid_, 0.0) + p.get('total_paid', 0)
        pd = (p.get('paid_at') or '')[:10]
        if pd and (not last_paid_by_stu.get(sid_) or pd > last_paid_by_stu[sid_]):
            last_paid_by_stu[sid_] = pd
        # A payment is late if made after the student's due_date.
        dd = due_date_by_stu.get(sid_)
        if dd and pd and pd > dd:
            late_flag_by_stu[sid_] = True

    # For monthly schedule computation we need per-student payments grouped
    payments_by_stu: Dict[str, list] = {}
    for p in payments:
        payments_by_stu.setdefault(p['student_id'], []).append(p)

    # Determine the current academic session — pick the most common one across
    # student assignments (fall back to 2026-27 if none).
    session_counts: Dict[str, int] = {}
    for a in assignments:
        s = a.get('academic_session')
        if s:
            session_counts[s] = session_counts.get(s, 0) + 1
    current_session = max(session_counts, key=session_counts.get) if session_counts else '2026-27'
    today_dt = datetime.now()

    today = datetime.now().date().isoformat()

    def _behavior_tag(paid: float, due: float, overdue_days: int, was_late: bool, expected: float) -> str:
        if expected <= 0:
            return 'na'
        if due <= 0 and paid > 0:
            return 'late' if was_late else 'regular'
        # Still owes something
        if overdue_days > 30:
            return 'defaulter'
        if was_late or overdue_days > 0:
            return 'late'
        return 'regular'

    rows = []
    for s in students:
        expected = expected_by_stu.get(s['id'], 0.0)
        disc = discount_by_stu.get(s['id'], 0.0)
        expected_after_disc = max(expected - disc, 0.0)
        paid = paid_by_stu.get(s['id'], 0.0)
        due = max(expected_after_disc - paid, 0.0)
        row_status = 'unpaid' if paid == 0 else ('paid' if due <= 0 else 'partial')
        last_paid = last_paid_by_stu.get(s['id'])
        dd = due_date_by_stu.get(s['id'])
        overdue_days = 0
        upcoming_due = None
        if dd:
            try:
                if dd < today and due > 0:
                    overdue_days = (datetime.fromisoformat(today) - datetime.fromisoformat(dd)).days
                elif dd >= today:
                    upcoming_due = dd
            except Exception:
                overdue_days = 0
        was_late = late_flag_by_stu.get(s['id'], False)
        behavior_tag = _behavior_tag(paid, due, overdue_days, was_late, expected_after_disc)
        collection_pct = round((paid / expected_after_disc) * 100, 1) if expected_after_disc > 0 else 0.0

        # Filter application
        if status_filter and status_filter != 'all' and status_filter != row_status:
            continue
        if due_min is not None and due < due_min:
            continue
        if due_max is not None and due > due_max:
            continue
        if payment_date_start and (not last_paid or last_paid < payment_date_start):
            continue
        if payment_date_end and (not last_paid or last_paid > payment_date_end):
            continue
        if behavior and behavior != 'all' and behavior_tag != behavior:
            continue
        if quick_view == 'defaulters' and behavior_tag != 'defaulter':
            continue
        if quick_view == 'fully_paid' and row_status != 'paid':
            continue
        if quick_view == 'upcoming' and not upcoming_due:
            continue

        rows.append({
            'student_id': s['id'],
            'admission_number': s.get('admission_number'),
            'full_name': s.get('full_name'),
            'class_id': s.get('class_id'),
            'class_name': class_map.get(s.get('class_id'), '-'),
            'section': s.get('section'),
            'phone': s.get('phone'),
            'father_name': s.get('father_name'),
            'expected': round(expected_after_disc, 2),
            'gross_expected': round(expected, 2),
            'discount': round(disc, 2),
            'paid': round(paid, 2),
            'due': round(due, 2),
            'collection_percent': collection_pct,
            'due_date': dd,
            'upcoming_due_date': upcoming_due,
            'last_payment_date': last_paid,
            'overdue_days': overdue_days,
            'status': row_status,
            'behavior_tag': behavior_tag,
        })

    # ---- Compute monthly schedule for every student in the filtered result ----
    for r in rows:
        sid_ = r['student_id']
        sched, monthly_amt, fully_paid, oc, oa, _due = _build_month_schedule(
            r['expected'], payments_by_stu.get(sid_, []), current_session, today_dt,
        )
        # Compact month tokens: only the fields the UI needs.
        r['monthly_amount'] = monthly_amt
        r['monthly_status'] = [
            {'i': e['index'], 'label': e['label'], 'status': e['status'],
             'paid': round(e['paid_amount'], 2), 'due': round(max(monthly_amt - e['paid_amount'], 0), 2)}
            for e in sched
        ]
        r['fully_paid'] = fully_paid
        r['overdue_months'] = oc
        r['overdue_amount'] = oa

    total_expected = sum(r['expected'] for r in rows)
    total_paid = sum(r['paid'] for r in rows)
    total_due = sum(r['due'] for r in rows)
    total_discount = sum(r['discount'] for r in rows)
    total_gross = sum(r['gross_expected'] for r in rows)
    collection_pct = round((total_paid / total_expected) * 100, 1) if total_expected > 0 else 0.0

    # Class/Section rollups — monthly-status oriented
    by_class_map: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        key = f"{r['class_id']}::{r.get('section') or '-'}"
        b = by_class_map.setdefault(key, {
            'class_id': r['class_id'], 'class_name': r['class_name'],
            'section': r.get('section') or '-',
            'students': 0, 'expected': 0.0, 'paid': 0.0, 'due': 0.0,
            'paid_count': 0, 'partial_count': 0, 'unpaid_count': 0,
            'fully_paid_count': 0, 'students_with_dues': 0,
            'overdue_amount': 0.0,
        })
        b['students'] += 1
        b['expected'] += r['expected']
        b['paid'] += r['paid']
        b['due'] += r['due']
        b[f"{r['status']}_count"] += 1
        if r.get('fully_paid'):
            b['fully_paid_count'] += 1
        if r['due'] > 0:
            b['students_with_dues'] += 1
        b['overdue_amount'] += r.get('overdue_amount', 0.0)
    by_class = sorted(by_class_map.values(), key=lambda x: (x['class_name'], x['section']))
    for b in by_class:
        b['collection_percent'] = round((b['paid'] / b['expected']) * 100, 1) if b['expected'] > 0 else 0.0
        b['overdue_amount'] = round(b['overdue_amount'], 2)

    return {
        'rows': rows,
        'count': len(rows),
        'by_class': by_class,
        'academic_session': current_session,
        'summary': {
            'total_expected': total_expected,
            'total_gross_expected': total_gross,
            'total_paid': total_paid,
            'total_due': total_due,
            'total_discount': total_discount,
            'collection_percent': collection_pct,
            'paid_count': sum(1 for r in rows if r['status'] == 'paid'),
            'partial_count': sum(1 for r in rows if r['status'] == 'partial'),
            'unpaid_count': sum(1 for r in rows if r['status'] == 'unpaid'),
            'fully_paid_count': sum(1 for r in rows if r.get('fully_paid')),
            'students_with_dues': sum(1 for r in rows if r['due'] > 0),
            'total_overdue_amount': round(sum(r.get('overdue_amount', 0) for r in rows), 2),
            'defaulter_count': sum(1 for r in rows if r['behavior_tag'] == 'defaulter'),
            'late_count': sum(1 for r in rows if r['behavior_tag'] == 'late'),
            'regular_count': sum(1 for r in rows if r['behavior_tag'] == 'regular'),
            'upcoming_count': sum(1 for r in rows if r['upcoming_due_date']),
        }
    }


# ------- Fee-status export endpoints (PDF / XLSX / CSV) -------

def _fee_status_export_columns() -> List[str]:
    return ['Admission No', 'Student', 'Class', 'Section', 'Guardian', 'Phone',
            'Expected (Rs.)', 'Discount (Rs.)', 'Paid (Rs.)', 'Due (Rs.)',
            'Due Date', 'Last Payment', 'Overdue Days', 'Status', 'Behavior']


def _fee_status_row_to_list(r: Dict[str, Any]) -> List[Any]:
    return [
        r.get('admission_number') or '',
        r.get('full_name') or '',
        r.get('class_name') or '',
        r.get('section') or '',
        r.get('father_name') or '',
        r.get('phone') or '',
        f"{r.get('expected', 0):,.2f}",
        f"{r.get('discount', 0):,.2f}",
        f"{r.get('paid', 0):,.2f}",
        f"{r.get('due', 0):,.2f}",
        r.get('due_date') or '',
        r.get('last_payment_date') or '',
        r.get('overdue_days') or 0,
        (r.get('status') or '').title(),
        (r.get('behavior_tag') or 'na').title(),
    ]


@api.get('/reports/monthly-dues.xlsx',
         dependencies=[Depends(require_roles('super_admin', 'school_admin', 'accountant'))])
async def report_monthly_dues_xlsx(request: Request,
                                   current=Depends(get_current_user),
                                   school_id: Optional[str] = None,
                                   class_sections: Optional[str] = None,
                                   only_with_dues: bool = True):
    """Export a **Due List** — for each student in the selected classes, list
    monthly-fee status (paid / partial / overdue / upcoming) across all 12
    months of the academic session, along with total due amount and
    student/guardian contact details. Ideal for classroom-level follow-ups.

    Query params:
      - `class_sections`: comma-separated `class_id:section` pairs (blank
        section = all sections). Omit to export all classes.
      - `only_with_dues` (default True): if True, only students who currently
        have Due > 0 are exported.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data = await report_fee_status(request, current, school_id, None, None,
                                   class_sections, None, None, None, None,
                                   None, None, None)
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    school = await schools_col.find_one({'id': sid}, {'_id': 0}) or {}

    rows = data['rows']
    if only_with_dues:
        rows = [r for r in rows if (r.get('due') or 0) > 0]
    # Sort by class → section → name for easy classroom distribution
    rows.sort(key=lambda r: (r.get('class_name') or '', r.get('section') or '',
                             r.get('full_name') or ''))

    # Month headers (short form: Apr, May, ...)
    month_short = []
    if rows:
        for m in rows[0].get('monthly_status') or []:
            lbl = m.get('label') or ''
            month_short.append(lbl.split(' ')[0][:3] if lbl else '')
    if not month_short:
        month_short = ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep',
                       'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Due List'

    # Title & meta
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8 + len(month_short))
    t = ws.cell(row=1, column=1, value=f"{school.get('name', 'Stanvard School')} — Fee Due List")
    t.font = Font(bold=True, size=14, color='0B2F4A')
    t.alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8 + len(month_short))
    m = ws.cell(row=2, column=1,
                value=f"Session {data.get('academic_session', '')} · {len(rows)} student(s) · "
                      f"Total due: Rs. {sum(r.get('due', 0) for r in rows):,.2f} · "
                      f"Generated on {datetime.now().strftime('%d %b %Y, %H:%M')}")
    m.font = Font(size=10, color='475569')
    m.alignment = Alignment(horizontal='center')

    header = ['Admission No', 'Student Name', 'Class', 'Section',
              'Guardian', 'Contact', 'Monthly (Rs.)'] + month_short + ['Total Due (Rs.)']
    ws.append([])  # blank row
    ws.append(header)
    header_row_idx = ws.max_row
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='0B2F4A')
    for col_idx in range(1, len(header) + 1):
        c = ws.cell(row=header_row_idx, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Color palette
    fills = {
        'paid_full':   PatternFill('solid', fgColor='0F766E'),   # deep green — fully paid
        'paid':        PatternFill('solid', fgColor='A7F3D0'),   # light green — that month paid
        'partial':     PatternFill('solid', fgColor='FDE68A'),   # amber
        'overdue':     PatternFill('solid', fgColor='FCA5A5'),   # red
        'pending':     PatternFill('solid', fgColor='F1F5F9'),   # grey
    }
    thin = Side(border_style='thin', color='CBD5E1')
    cell_border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for r in rows:
        ws.append([
            r.get('admission_number') or '',
            r.get('full_name') or '',
            r.get('class_name') or '',
            r.get('section') or '',
            r.get('father_name') or '',
            r.get('phone') or '',
            round(r.get('monthly_amount') or 0, 2),
        ] + [
            # Cell text: '✓' for paid, amount for partial/overdue, blank for pending
            (
                '✓' if m['status'] == 'paid' else
                round(m['due'], 0) if m['status'] in ('overdue', 'partial') else
                ''
            )
            for m in (r.get('monthly_status') or [])
        ] + [round(r.get('due') or 0, 2)])

        row_idx = ws.max_row
        # Colour the month cells
        for j, m in enumerate((r.get('monthly_status') or [])):
            cell = ws.cell(row=row_idx, column=8 + j)  # 7 leading cols + 1 (1-indexed)
            key = 'paid_full' if r.get('fully_paid') and m['status'] == 'paid' else m['status']
            cell.fill = fills.get(key, fills['pending'])
            cell.alignment = Alignment(horizontal='center')
            if key == 'paid_full':
                cell.font = Font(bold=True, color='FFFFFF')
            elif m['status'] == 'overdue':
                cell.font = Font(bold=True, color='B42318')
        # Border for whole row
        for col_idx in range(1, len(header) + 1):
            ws.cell(row=row_idx, column=col_idx).border = cell_border
        # Bold the last "Total Due" cell
        last = ws.cell(row=row_idx, column=len(header))
        last.font = Font(bold=True, color='B42318')
        last.alignment = Alignment(horizontal='right')

    # Column widths
    widths = [14, 30, 12, 8, 22, 14, 12] + [7] * len(month_short) + [14]
    from openpyxl.utils import get_column_letter
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.row_dimensions[header_row_idx].height = 26
    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=3)

    # Legend on a second sheet
    leg = wb.create_sheet('Legend')
    leg.append(['Colour', 'Meaning'])
    for c in leg[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center')
    legend_rows = [
        ('Fully paid (all 12 months)', 'paid_full'),
        ('Monthly fee paid', 'paid'),
        ('Partial payment', 'partial'),
        ('Overdue (past due date)', 'overdue'),
        ('Upcoming month (not yet due)', 'pending'),
    ]
    for label, key in legend_rows:
        leg.append([' ', label])
        cell = leg.cell(row=leg.max_row, column=1)
        cell.fill = fills[key]
        cell.border = cell_border
    leg.column_dimensions['A'].width = 12
    leg.column_dimensions['B'].width = 36

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = 'due_list.xlsx'
    return StreamingResponse(
        buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )


# Shared query args for all fee-status export endpoints
_FS_QP = dict(
    school_id=None, class_id=None, section=None, class_sections=None,
    status_filter=None, due_min=None, due_max=None,
    payment_date_start=None, payment_date_end=None,
    quick_view=None, behavior=None,
)


@api.get('/reports/fee-status.pdf', dependencies=[Depends(require_roles('super_admin', 'school_admin', 'accountant'))])
async def report_fee_status_pdf(request: Request, current=Depends(get_current_user),
                                school_id: Optional[str] = None,
                                class_id: Optional[str] = None,
                                section: Optional[str] = None,
                                class_sections: Optional[str] = None,
                                status_filter: Optional[str] = None,
                                due_min: Optional[float] = None,
                                due_max: Optional[float] = None,
                                payment_date_start: Optional[str] = None,
                                payment_date_end: Optional[str] = None,
                                quick_view: Optional[str] = None,
                                behavior: Optional[str] = None):
    data = await report_fee_status(request, current, school_id, class_id, section, class_sections,
                                   status_filter, due_min, due_max, payment_date_start, payment_date_end,
                                   quick_view, behavior)
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    school = await schools_col.find_one({'id': sid}, {'_id': 0}) or {}
    rows = [_fee_status_row_to_list(r) for r in data['rows']]
    subtitle_bits: List[str] = []
    if class_sections:
        subtitle_bits.append(f"Classes/Sections: {class_sections}")
    if status_filter and status_filter != 'all':
        subtitle_bits.append(f"Status: {status_filter}")
    if quick_view:
        subtitle_bits.append(f"View: {quick_view}")
    if behavior and behavior != 'all':
        subtitle_bits.append(f"Behavior: {behavior}")
    subtitle = ' | '.join(subtitle_bits) if subtitle_bits else 'All Classes / Sections'
    summary = {
        'Students': data['count'],
        'Total Expected': f"Rs. {data['summary']['total_expected']:,.2f}",
        'Total Paid': f"Rs. {data['summary']['total_paid']:,.2f}",
        'Total Due': f"Rs. {data['summary']['total_due']:,.2f}",
        'Collection %': f"{data['summary']['collection_percent']}%",
        'Paid / Partial / Unpaid': (
            f"{data['summary']['paid_count']} / {data['summary']['partial_count']} / {data['summary']['unpaid_count']}"
        ),
        'Defaulters': data['summary']['defaulter_count'],
    }
    pdf = generate_report_pdf('Student Fee Status Report', subtitle,
                              _fee_status_export_columns(), rows,
                              school.get('name', 'Stanvard School'), summary)
    return StreamingResponse(io.BytesIO(pdf), media_type='application/pdf',
                             headers={'Content-Disposition': 'inline; filename="fee_status_report.pdf"'})


@api.get('/reports/fee-status.xlsx', dependencies=[Depends(require_roles('super_admin', 'school_admin', 'accountant'))])
async def report_fee_status_xlsx(request: Request, current=Depends(get_current_user),
                                 school_id: Optional[str] = None,
                                 class_id: Optional[str] = None,
                                 section: Optional[str] = None,
                                 class_sections: Optional[str] = None,
                                 status_filter: Optional[str] = None,
                                 due_min: Optional[float] = None,
                                 due_max: Optional[float] = None,
                                 payment_date_start: Optional[str] = None,
                                 payment_date_end: Optional[str] = None,
                                 quick_view: Optional[str] = None,
                                 behavior: Optional[str] = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    data = await report_fee_status(request, current, school_id, class_id, section, class_sections,
                                   status_filter, due_min, due_max, payment_date_start, payment_date_end,
                                   quick_view, behavior)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Fee Status'
    header = _fee_status_export_columns()
    ws.append(header)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0B2F4A')
    for col_idx, _ in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for r in data['rows']:
        ws.append([
            r.get('admission_number') or '',
            r.get('full_name') or '',
            r.get('class_name') or '',
            r.get('section') or '',
            r.get('father_name') or '',
            r.get('phone') or '',
            r.get('expected', 0),
            r.get('discount', 0),
            r.get('paid', 0),
            r.get('due', 0),
            r.get('due_date') or '',
            r.get('last_payment_date') or '',
            r.get('overdue_days') or 0,
            (r.get('status') or '').title(),
            (r.get('behavior_tag') or 'na').title(),
        ])
    widths = [16, 26, 14, 10, 22, 14, 14, 14, 14, 14, 12, 14, 12, 12, 12]
    for i, w in enumerate(widths, 1):
        # Use openpyxl utils to handle >26 columns safely
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w
    s = data['summary']
    ws2 = wb.create_sheet('Summary')
    ws2.append(['Metric', 'Value'])
    for k, v in {
        'Total Students': data['count'],
        'Total Expected (Rs.)': s['total_expected'],
        'Total Paid (Rs.)': s['total_paid'],
        'Total Due (Rs.)': s['total_due'],
        'Total Discount (Rs.)': s['total_discount'],
        'Collection %': s['collection_percent'],
        'Paid Students': s['paid_count'],
        'Partial Students': s['partial_count'],
        'Unpaid Students': s['unpaid_count'],
        'Defaulters': s['defaulter_count'],
        'Late Payers': s['late_count'],
        'Regular Payers': s['regular_count'],
    }.items():
        ws2.append([k, v])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 20

    # Class rollup sheet
    ws3 = wb.create_sheet('By Class')
    ws3.append(['Class', 'Section', 'Students', 'Expected', 'Paid', 'Due', 'Collection %',
                'Paid', 'Partial', 'Unpaid', 'Defaulters'])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
    for b in data['by_class']:
        ws3.append([b['class_name'], b['section'], b['students'],
                    b['expected'], b['paid'], b['due'], b['collection_percent'],
                    b['paid_count'], b['partial_count'], b['unpaid_count'], b['defaulters']])
    for col, w in enumerate([14, 10, 10, 14, 14, 14, 12, 8, 8, 8, 10], 1):
        from openpyxl.utils import get_column_letter
        ws3.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                             headers={'Content-Disposition': 'attachment; filename="fee_status_report.xlsx"'})


@api.get('/reports/fee-status.csv', dependencies=[Depends(require_roles('super_admin', 'school_admin', 'accountant'))])
async def report_fee_status_csv(request: Request, current=Depends(get_current_user),
                                school_id: Optional[str] = None,
                                class_id: Optional[str] = None,
                                section: Optional[str] = None,
                                class_sections: Optional[str] = None,
                                status_filter: Optional[str] = None,
                                due_min: Optional[float] = None,
                                due_max: Optional[float] = None,
                                payment_date_start: Optional[str] = None,
                                payment_date_end: Optional[str] = None,
                                quick_view: Optional[str] = None,
                                behavior: Optional[str] = None):
    import csv
    data = await report_fee_status(request, current, school_id, class_id, section, class_sections,
                                   status_filter, due_min, due_max, payment_date_start, payment_date_end,
                                   quick_view, behavior)
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(_fee_status_export_columns())
    for r in data['rows']:
        writer.writerow(_fee_status_row_to_list(r))
    return StreamingResponse(io.BytesIO(buffer.getvalue().encode()),
                             media_type='text/csv',
                             headers={'Content-Disposition': 'attachment; filename="fee_status_report.csv"'})


# =====================================================
# FEE ANALYTICS (dedicated, filterable)
# =====================================================
@api.get('/analytics/fees')
async def fee_analytics(request: Request, current=Depends(get_current_user),
                        school_id: Optional[str] = None,
                        start_date: Optional[str] = None,
                        end_date: Optional[str] = None,
                        class_id: Optional[str] = None,
                        section: Optional[str] = None,
                        payment_mode: Optional[str] = None,
                        payment_status: Optional[str] = None):
    sid = resolve_school_id(current, school_id, request.headers.get('X-School-Id'))
    today_iso = datetime.now().strftime('%Y-%m-%d')

    # Payment query
    pq: Dict[str, Any] = {'school_id': sid, 'status': 'success'}
    if start_date or end_date:
        rq: Dict[str, Any] = {}
        if start_date:
            rq['$gte'] = start_date
        if end_date:
            rq['$lte'] = end_date + 'T23:59:59'
        pq['paid_at'] = rq
    if payment_mode:
        pq['payment_mode'] = payment_mode
    payments_all = await payments_col.find(pq, {'_id': 0}).to_list(50000)

    # Filter by class/section (need student data)
    students_all = await students_col.find({'school_id': sid}, {'_id': 0}).to_list(20000)
    stu_by_id = {s['id']: s for s in students_all}
    def _in_scope(pmt):
        s = stu_by_id.get(pmt.get('student_id'))
        if not s:
            return False
        if class_id and s.get('class_id') != class_id:
            return False
        if section and s.get('section') != section:
            return False
        return True
    payments = [p for p in payments_all if _in_scope(p)]

    total_collected = sum(p.get('total_paid', 0) for p in payments)
    total_discount = sum(p.get('discount', 0) for p in payments)
    total_late_fee = sum(p.get('late_fee', 0) for p in payments)

    # Today / month specific (independent of the filter for context KPIs)
    month_start = datetime.now().strftime('%Y-%m-01')
    today_payments = await payments_col.find({'school_id': sid, 'status': 'success',
                                              'paid_at': {'$gte': today_iso, '$lte': today_iso + 'T23:59:59'}}, {'_id': 0}).to_list(5000)
    month_payments = await payments_col.find({'school_id': sid, 'status': 'success',
                                              'paid_at': {'$gte': month_start}}, {'_id': 0}).to_list(20000)

    # Expected & pending computation (fee-status logic, filtered by class/section)
    students_scope = [s for s in students_all if s.get('status') == 'active'
                      and (not class_id or s.get('class_id') == class_id)
                      and (not section or s.get('section') == section)]
    student_ids_scope = [s['id'] for s in students_scope]
    assignments = await fee_assignments_col.find({'student_id': {'$in': student_ids_scope}}, {'_id': 0}).to_list(50000)
    plan_cache: Dict[str, Any] = {}
    async def _get_plan(pid):
        if pid not in plan_cache:
            plan_cache[pid] = await fee_plans_col.find_one({'id': pid}, {'_id': 0}) or {}
        return plan_cache[pid]

    expected_by_stu: Dict[str, float] = {}
    discount_by_stu: Dict[str, float] = {}
    for a in assignments:
        items = a.get('custom_items') or []
        if not items and a.get('fee_plan_id'):
            plan = await _get_plan(a['fee_plan_id'])
            items = plan.get('items', [])
        expected_by_stu[a['student_id']] = expected_by_stu.get(a['student_id'], 0.0) + sum(i.get('amount', 0) for i in items)
        discount_by_stu[a['student_id']] = discount_by_stu.get(a['student_id'], 0.0) + a.get('discount_amount', 0)

    paid_by_stu: Dict[str, float] = {}
    all_paid_by_stu = await payments_col.find({'student_id': {'$in': student_ids_scope}, 'status': 'success'}, {'_id': 0}).to_list(50000)
    for p in all_paid_by_stu:
        paid_by_stu[p['student_id']] = paid_by_stu.get(p['student_id'], 0.0) + p.get('total_paid', 0)

    paid_count = 0
    partial_count = 0
    unpaid_count = 0
    total_expected = 0.0
    total_pending = 0.0
    for s in students_scope:
        exp = expected_by_stu.get(s['id'], 0.0)
        disc = discount_by_stu.get(s['id'], 0.0)
        pd = paid_by_stu.get(s['id'], 0.0)
        due = max(exp - disc - pd, 0.0)
        total_expected += exp
        total_pending += due
        if pd <= 0:
            unpaid_count += 1
        elif due <= 0:
            paid_count += 1
        else:
            partial_count += 1

    # Apply payment_status filter to KPIs if requested (only for pending/paid student counts)
    # (already computed all; payment_status just narrows which student cards are displayed - handled client-side)

    # Daily buckets across the filter range (or last 14 days if no range)
    if start_date and end_date:
        try:
            d0 = datetime.strptime(start_date, '%Y-%m-%d')
            d1 = datetime.strptime(end_date, '%Y-%m-%d')
        except Exception:
            d0 = datetime.now() - timedelta(days=13)
            d1 = datetime.now()
    else:
        d0 = datetime.now() - timedelta(days=13)
        d1 = datetime.now()
    days = min((d1 - d0).days + 1, 92)  # cap at ~3 months for the daily view
    daily = []
    for i in range(days):
        d = (d0 + timedelta(days=i)).strftime('%Y-%m-%d')
        amt = sum(p.get('total_paid', 0) for p in payments if (p.get('paid_at') or '').startswith(d))
        cnt = sum(1 for p in payments if (p.get('paid_at') or '').startswith(d))
        daily.append({'date': d, 'amount': amt, 'transactions': cnt})

    # Monthly (12 months of current year - independent of filter)
    year = datetime.now().year
    monthly = []
    for m in range(1, 13):
        m_start = f'{year}-{m:02d}-01'
        m_end = f'{year}-{m + 1:02d}-01' if m < 12 else f'{year + 1}-01-01'
        mpays = [p for p in payments_all if m_start <= (p.get('paid_at') or '') < m_end]
        monthly.append({
            'month': datetime(year, m, 1).strftime('%b'),
            'amount': sum(p.get('total_paid', 0) for p in mpays),
            'transactions': len(mpays),
        })

    # Payment mode breakdown
    by_mode: Dict[str, Dict[str, Any]] = {}
    for p in payments:
        m = p.get('payment_mode', 'cash')
        b = by_mode.setdefault(m, {'amount': 0.0, 'count': 0})
        b['amount'] += p.get('total_paid', 0)
        b['count'] += 1

    # Class-wise collection
    classes = await classes_col.find({'school_id': sid}, {'_id': 0}).to_list(200)
    cls_map = {c['id']: c['name'] for c in classes}
    by_class_map: Dict[str, Dict[str, Any]] = {}
    for p in payments:
        s = stu_by_id.get(p.get('student_id'))
        if not s:
            continue
        cname = cls_map.get(s.get('class_id'), '—')
        b = by_class_map.setdefault(cname, {'amount': 0.0, 'transactions': 0, 'students': set()})
        b['amount'] += p.get('total_paid', 0)
        b['transactions'] += 1
        b['students'].add(s['id'])
    by_class = sorted(
        [{'class_name': k, 'amount': v['amount'], 'transactions': v['transactions'],
          'students': len(v['students'])} for k, v in by_class_map.items()],
        key=lambda x: x['amount'], reverse=True,
    )

    # Transactions list (enriched, sorted by paid_at desc) for the currently applied filter
    transactions = []
    for p in payments:
        s = stu_by_id.get(p.get('student_id')) or {}
        items = p.get('items') or []
        # Compose a compact fee-heads label (e.g., "Tuition, Transport")
        fee_heads = ', '.join(
            [i.get('fee_head_name') or '' for i in items if i.get('fee_head_name')]
        ) or '—'
        transactions.append({
            'id': p.get('id'),
            'receipt_number': p.get('receipt_number') or '—',
            'paid_at': p.get('paid_at'),
            'student_id': p.get('student_id'),
            'student_name': p.get('student_name') or s.get('full_name') or '—',
            'admission_number': s.get('admission_number') or '—',
            'class_name': cls_map.get(s.get('class_id'), '—'),
            'section': s.get('section') or '—',
            'father_name': s.get('father_name') or '—',
            'phone': s.get('phone') or '—',
            'fee_heads': fee_heads,
            'subtotal': p.get('subtotal', 0),
            'discount': p.get('discount', 0),
            'late_fee': p.get('late_fee', 0),
            'total_paid': p.get('total_paid', 0),
            'payment_mode': p.get('payment_mode', 'cash'),
            'txn_ref': p.get('txn_ref') or '',
            'status': p.get('status', 'success'),
            'collected_by_name': p.get('collected_by_name') or '—',
            'remarks': p.get('remarks') or '',
        })
    # Sort newest first
    transactions.sort(key=lambda t: (t.get('paid_at') or ''), reverse=True)

    return {
        'kpis': {
            'total_collected': total_collected,
            'total_pending': total_pending,
            'total_expected': total_expected,
            'total_paid_students': paid_count,
            'total_partial_students': partial_count,
            'total_pending_students': unpaid_count,
            'today_collection': sum(p.get('total_paid', 0) for p in today_payments),
            'today_transactions': len(today_payments),
            'monthly_collection': sum(p.get('total_paid', 0) for p in month_payments),
            'monthly_transactions': len(month_payments),
            'total_discount': total_discount,
            'total_late_fee': total_late_fee,
            'transactions_in_range': len(payments),
        },
        'daily': daily,
        'monthly': monthly,
        'by_mode': by_mode,
        'by_class': by_class,
        'transactions': transactions,
        'range': {'start_date': start_date, 'end_date': end_date},
    }


@api.get('/analytics/student/{student_id}/fee-report')
async def student_fee_report(student_id: str, request: Request, current=Depends(get_current_user)):
    """Full fee profile for a single student."""
    student = await students_col.find_one({'id': student_id}, {'_id': 0})
    if not student:
        raise HTTPException(404, 'Student not found')
    if current['role'] != 'super_admin' and student['school_id'] != current.get('school_id'):
        raise HTTPException(403, 'Forbidden')
    if current['role'] == 'parent' and not parent_can_access_student(current, student_id):
        raise HTTPException(403, 'Forbidden')

    assignments = await fee_assignments_col.find({'student_id': student_id}, {'_id': 0}).to_list(50)
    plan_cache: Dict[str, Any] = {}
    total_expected = 0.0
    total_discount = 0.0
    due_dates: List[str] = []
    line_items: List[Dict[str, Any]] = []
    for a in assignments:
        items = a.get('custom_items') or []
        if not items and a.get('fee_plan_id'):
            if a['fee_plan_id'] not in plan_cache:
                plan_cache[a['fee_plan_id']] = await fee_plans_col.find_one({'id': a['fee_plan_id']}, {'_id': 0}) or {}
            items = plan_cache[a['fee_plan_id']].get('items', [])
        for it in items:
            line_items.append({
                'fee_head_name': it.get('fee_head_name'),
                'frequency': it.get('frequency', 'monthly'),
                'amount': it.get('amount', 0),
                'due_date': it.get('due_date') or a.get('due_date'),
            })
            total_expected += it.get('amount', 0)
        total_discount += a.get('discount_amount', 0)
        if a.get('due_date'):
            due_dates.append(a['due_date'])

    payments = await payments_col.find({'student_id': student_id, 'status': 'success'}, {'_id': 0}).sort('paid_at', -1).to_list(500)
    total_paid = sum(p.get('total_paid', 0) for p in payments)
    total_late_paid = sum(p.get('late_fee', 0) for p in payments)
    balance = max(total_expected - total_discount - total_paid, 0)

    today = datetime.now().date()
    next_due_date = None
    days_overdue = 0
    if due_dates:
        due_sorted = sorted(due_dates)
        next_due_date = due_sorted[0]
        try:
            dd = datetime.strptime(next_due_date, '%Y-%m-%d').date()
            if dd < today and balance > 0:
                days_overdue = (today - dd).days
        except Exception:
            pass

    status = 'unpaid' if total_paid == 0 else ('paid' if balance <= 0 else 'partial')
    last_payment_date = payments[0]['paid_at'][:10] if payments else None

    # Class name
    class_name = '—'
    if student.get('class_id'):
        c = await classes_col.find_one({'id': student['class_id']}, {'_id': 0})
        if c:
            class_name = c['name']

    return {
        'student': {**student, 'class_name': class_name},
        'summary': {
            'total_expected': total_expected,
            'total_discount': total_discount,
            'total_paid': total_paid,
            'total_late_paid': total_late_paid,
            'balance': balance,
            'status': status,
            'last_payment_date': last_payment_date,
            'next_due_date': next_due_date,
            'days_overdue': days_overdue,
        },
        'line_items': line_items,
        'payments': payments,
    }


@api.get('/analytics/student/{student_id}/fee-report.pdf')
async def student_fee_report_pdf(student_id: str, request: Request, current=Depends(get_current_user)):
    data = await student_fee_report(student_id, request, current)
    school = await schools_col.find_one({'id': data['student']['school_id']}, {'_id': 0}) or {}
    s = data['student']
    summary = data['summary']

    cols = ['Receipt No', 'Date', 'Mode', 'Late Fee', 'Discount', 'Amount (Rs.)']
    rows = []
    for p in data['payments']:
        rows.append([
            p.get('receipt_number', ''),
            (p.get('paid_at') or '')[:10],
            str(p.get('payment_mode', '')).replace('_', ' ').title(),
            f"{p.get('late_fee', 0):,.2f}",
            f"{p.get('discount', 0):,.2f}",
            f"{p.get('total_paid', 0):,.2f}",
        ])
    subtitle = (f"Student Fee Report — {s.get('full_name')} ({s.get('admission_number')}) — "
               f"Class {s.get('class_name')} {s.get('section') or ''}")
    summary_line = {
        'Total Fee': f"Rs. {summary['total_expected']:,.2f}",
        'Total Paid': f"Rs. {summary['total_paid']:,.2f}",
        'Discount': f"Rs. {summary['total_discount']:,.2f}",
        'Late Fee': f"Rs. {summary['total_late_paid']:,.2f}",
        'Balance': f"Rs. {summary['balance']:,.2f}",
        'Status': summary['status'].upper(),
    }
    if summary.get('next_due_date'):
        summary_line['Next Due'] = summary['next_due_date']
    pdf = generate_report_pdf('Student Fee Report', subtitle, cols, rows,
                              school.get('name', 'Stanvard School'), summary_line)
    fname = f'fee_report_{s.get("admission_number", student_id)}.pdf'
    return StreamingResponse(io.BytesIO(pdf), media_type='application/pdf',
                             headers={'Content-Disposition': f'inline; filename="{fname}"'})


# ------------ Mount router & middlewares ------------
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=['*'],
    allow_headers=['*'],
)


@app.on_event('startup')
async def startup():
    # Create indexes
    await users_col.create_index('email', unique=True)
    await students_col.create_index([('school_id', 1), ('admission_number', 1)])
    await students_col.create_index([('school_id', 1), ('class_id', 1)])
    await payments_col.create_index([('school_id', 1), ('paid_at', -1)])
    await attendance_col.create_index([('school_id', 1), ('date', 1), ('class_id', 1)])
    logger.info('Stanvard ERP API started')


@app.on_event('shutdown')
async def shutdown():
    mongo_client.close()
